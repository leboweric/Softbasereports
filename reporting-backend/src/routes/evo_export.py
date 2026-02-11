"""
EVO Export Route - Tenant-specific P&L Excel export
Populates the tenant's EVO template with GL data from the IPS database.
Pre-computes all VLOOKUP formula results server-side so numbers display
immediately when opened in any version of Excel.

Uses openpyxl to load/save the template, matching Amy's proven approach.
"""

from flask import Blueprint, jsonify, request, send_file
from datetime import datetime
import logging
import re
import os
from io import BytesIO
from openpyxl import load_workbook
from flask_jwt_extended import jwt_required, get_jwt_identity
from src.utils.tenant_utils import get_tenant_db
from src.models.user import User

logger = logging.getLogger(__name__)
evo_export_bp = Blueprint('evo_export', __name__)

# Map organization schema to template file
TEMPLATE_MAP = {
    'ind004': 'IPS_template.xlsx',
    # Add more tenants here as needed:
    # 'bmh001': 'BMH_template.xlsx',
}


def get_tenant_schema():
    """Get the database schema for the current user's organization"""
    try:
        user_id = get_jwt_identity()
        if user_id:
            user = User.query.get(int(user_id))
            if user and user.organization and user.organization.database_schema:
                return user.organization.database_schema
        return 'ben002'
    except:
        return 'ben002'


def get_tenant_db_service():
    return get_tenant_db()


def get_all_gl_data(schema, year, month):
    """
    Query ALL GL accounts for a given year/month.
    Returns a list of dicts with AccountNo, Year, Month, YTD, MTD, Description, Type.
    """
    query = f"""
    SELECT 
        g.AccountNo,
        g.Year,
        g.Month,
        g.YTD,
        g.MTD,
        c.Description,
        c.Type
    FROM {schema}.GL g
    LEFT JOIN {schema}.ChartOfAccounts c ON g.AccountNo = c.AccountNo
    WHERE g.Year = %s AND g.Month = %s
    ORDER BY g.AccountNo
    """
    try:
        results = get_tenant_db_service().execute_query(query, [year, month])
        rows = []
        for r in results:
            rows.append({
                'AccountNo': str(r.get('AccountNo', '')).strip(),
                'Year': int(r.get('Year', year)),
                'Month': int(r.get('Month', month)),
                'YTD': float(r.get('YTD', 0) or 0),
                'MTD': float(r.get('MTD', 0) or 0),
                'Description': str(r.get('Description', '') or ''),
                'Type': str(r.get('Type', '') or ''),
            })
        return rows
    except Exception as e:
        logger.error(f"Error fetching GL data for {schema} {year}-{month}: {e}")
        return []


def get_prior_month(year, month):
    """Return (year, month) for the prior month"""
    if month == 1:
        return (year - 1, 12)
    return (year, month - 1)


def build_lookup_dicts(current_data, prior_year_data, prior_month_data):
    """
    Build lookup dictionaries keyed by account number string.
    Returns (tb_lookup, tb1_lookup, tb2_lookup)
    """
    tb_lookup = {}
    for acct in current_data:
        tb_lookup[acct['AccountNo']] = acct

    tb1_lookup = {}
    for acct in prior_year_data:
        tb1_lookup[acct['AccountNo']] = acct

    tb2_lookup = {}
    for acct in prior_month_data:
        tb2_lookup[acct['AccountNo']] = acct

    return tb_lookup, tb1_lookup, tb2_lookup


# VLOOKUP column index to data field mapping
TB_COL_MAP = {1: 'AccountNo', 2: 'Year', 3: 'Month', 4: 'YTD', 5: 'MTD', 6: 'Description', 7: 'Type'}
TB2_COL_MAP = {1: 'AccountNo', 2: 'AccountField', 3: 'YTD', 4: 'MTD', 5: 'Type', 6: 'Description'}

# Regex to parse VLOOKUP formulas
VLOOKUP_RE = re.compile(
    r'^=VLOOKUP\(([A-Z]+)(\d+),TB!(TB|TB1_1|TB2_),(\d+),FALSE\)(?:\*(-?1))?$'
)


def resolve_vlookup(formula, sheet, tb_lookup, tb1_lookup, tb2_lookup):
    """
    Parse a VLOOKUP formula and compute its result.
    Returns the computed value, or 0 if the account is not found.
    """
    match = VLOOKUP_RE.match(formula)
    if not match:
        return None

    ref_row = int(match.group(2))
    table_name = match.group(3)
    col_idx = int(match.group(4))
    multiplier_str = match.group(5)
    multiplier = int(multiplier_str) if multiplier_str else 1

    # Get the lookup value (account number) from column A of the referenced row
    lookup_value = sheet.cell(row=ref_row, column=1).value
    if lookup_value is None:
        return 0

    # Convert to string for lookup
    if isinstance(lookup_value, (int, float)):
        lookup_key = str(int(lookup_value))
    else:
        lookup_key = str(lookup_value).strip()

    # Select the right lookup dict and column map
    if table_name == 'TB':
        lookup_dict = tb_lookup
        col_map = TB_COL_MAP
    elif table_name == 'TB1_1':
        lookup_dict = tb1_lookup
        col_map = TB_COL_MAP
    elif table_name == 'TB2_':
        lookup_dict = tb2_lookup
        col_map = TB2_COL_MAP
    else:
        return 0

    acct_data = lookup_dict.get(lookup_key)
    if acct_data is None:
        return 0

    field_name = col_map.get(col_idx)
    if field_name is None:
        return 0

    value = acct_data.get(field_name, 0)
    if value is None:
        value = 0

    if isinstance(value, (int, float)):
        return value * multiplier
    return value


def precompute_all_vlookups(wb, tb_lookup, tb1_lookup, tb2_lookup):
    """
    Efficiently scan only the specific columns known to contain VLOOKUP formulas
    in each sheet. This avoids scanning all 16,384 columns which causes timeouts.
    """
    total_computed = 0

    # Known VLOOKUP column locations per sheet (from template analysis)
    # Format: {sheet_name: (min_row, max_row, [columns])}
    VLOOKUP_LOCATIONS = {
        'Balance Sheet':               (5, 175, [5, 7]),
        'Trial Balance':               (7, 130, [5, 7, 9, 13]),
        'Combined Detail P and L':     (7, 638, [5, 8, 11, 14]),
        'Dynamic Storage Solutions':   (6, 59,  [5, 8, 11, 14]),
        'C H Steel Solutions':         (6, 60,  [5, 8, 11, 14]),
        'AMI Sales Department':        (7, 52,  [5, 8, 11, 14]),
        'Canton Sales Department':     (7, 59,  [5, 8, 11, 14]),
        'Canton Parts Department':     (7, 85,  [5, 8, 11, 14]),
        'Canton Service Department':   (7, 80,  [7, 10, 13, 16]),
        'Canton Rental Department':    (7, 39,  [5, 8, 11, 14]),
        'Canton Used Department':      (7, 48,  [5, 8, 11, 14]),
        'Administration Department':   (7, 51,  [5, 8, 11, 14]),
        'Cleveland Sales Department':  (7, 62,  [5, 8, 11, 14]),
        'Cleveland Parts Department':  (7, 76,  [5, 8, 11, 14]),
        'Cleveland Service Department':(7, 80,  [7, 10, 13, 16]),
        'Cleveland Rental Department': (7, 37,  [5, 8, 11, 14]),
        'Cleveland Used Department':   (7, 48,  [5, 8, 11, 14]),
    }

    for sheet_name in wb.sheetnames:
        if sheet_name == 'TB':
            continue

        ws = wb[sheet_name]
        sheet_computed = 0

        if sheet_name in VLOOKUP_LOCATIONS:
            # Fast path: only scan known VLOOKUP cells
            min_row, max_row, cols = VLOOKUP_LOCATIONS[sheet_name]
            for row_idx in range(min_row, max_row + 1):
                for col_idx in cols:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('=VLOOKUP'):
                        result = resolve_vlookup(cell.value, ws, tb_lookup, tb1_lookup, tb2_lookup)
                        if result is not None:
                            cell.value = result
                            sheet_computed += 1
        else:
            # Fallback: scan all cells but limit to max_col=20 to avoid 16384 col issue
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 700), max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('=VLOOKUP'):
                        result = resolve_vlookup(cell.value, ws, tb_lookup, tb1_lookup, tb2_lookup)
                        if result is not None:
                            cell.value = result
                            sheet_computed += 1

        if sheet_computed > 0:
            logger.info(f"  Pre-computed {sheet_computed} VLOOKUPs in '{sheet_name}'")
            total_computed += sheet_computed

    logger.info(f"Total VLOOKUPs pre-computed: {total_computed}")
    return total_computed


@evo_export_bp.route('/api/reports/pl/evo/export', methods=['GET'])
@jwt_required()
def export_evo():
    """
    Export P&L using the tenant-specific EVO Excel template.
    
    The template has a TB sheet with 3 Excel tables:
    - Table_TB (B5:H761): Current year/month data
    - Table_TB1_1 (P5:V753): Prior year same month data
    - Table_TB2_ (Y5:AD752): Prior month data (beginning balances)
    
    All other sheets have VLOOKUP formulas referencing these tables.
    We pre-compute all VLOOKUP results so numbers display immediately.
    
    Uses openpyxl load/save (same approach as Amy's proven working code).
    
    Query Parameters:
        year: Year for the report (default: current year)
        month: Month for the report (default: current month)
    """
    try:
        schema = get_tenant_schema()
        
        # Get parameters
        now = datetime.now()
        year = request.args.get('year', type=int, default=now.year)
        month = request.args.get('month', type=int, default=now.month)
        
        logger.info(f"Generating EVO export for {year}-{month:02d}, schema: {schema}")
        
        # Find the template for this tenant
        template_file = TEMPLATE_MAP.get(schema)
        if not template_file:
            return jsonify({
                'error': f'No EVO template configured for this organization ({schema}). '
                         f'Please contact support to set up your custom template.'
            }), 404
        
        template_path = os.path.join(
            os.path.dirname(__file__), '..', 'templates', 'evo', template_file
        )
        
        if not os.path.exists(template_path):
            return jsonify({
                'error': f'Template file not found: {template_file}'
            }), 500
        
        # Load the template workbook (preserving formulas)
        wb = load_workbook(template_path)
        ws = wb['TB']
        
        # --- Fetch GL data for 3 periods ---
        
        # 1. Current year/month
        current_data = get_all_gl_data(schema, year, month)
        logger.info(f"Current period ({year}-{month}): {len(current_data)} accounts")
        
        # 2. Prior year, same month
        prior_year_data = get_all_gl_data(schema, year - 1, month)
        logger.info(f"Prior year ({year-1}-{month}): {len(prior_year_data)} accounts")
        
        # 3. Prior month (for beginning balances)
        prev_year, prev_month = get_prior_month(year, month)
        prior_month_data = get_all_gl_data(schema, prev_year, prev_month)
        logger.info(f"Prior month ({prev_year}-{prev_month}): {len(prior_month_data)} accounts")
        
        # --- Update control cells ---
        ws.cell(row=3, column=1, value=month)
        ws.cell(row=4, column=1, value=year)
        
        # --- Populate Table_TB (columns B-H, starting row 6) ---
        # Clear existing data first
        for row in range(6, 761):  # Don't clear row 761 (totals row)
            for col in [2, 3, 4, 5, 6, 7, 8]:
                ws.cell(row=row, column=col, value=None)
        
        for i, acct in enumerate(current_data):
            row = 6 + i
            if row > 760:  # Stop before totals row
                break
            ws.cell(row=row, column=2, value=int(acct['AccountNo']) if acct['AccountNo'].isdigit() else acct['AccountNo'])
            ws.cell(row=row, column=3, value=acct['Year'])
            ws.cell(row=row, column=4, value=acct['Month'])
            ws.cell(row=row, column=5, value=acct['YTD'])
            ws.cell(row=row, column=6, value=acct['MTD'])
            ws.cell(row=row, column=7, value=acct['Description'])
            ws.cell(row=row, column=8, value=acct['Type'])
        
        # --- Populate Table_TB1_1 (columns P-V, starting row 6) ---
        for row in range(6, 753):  # Don't clear totals row
            for col in [16, 17, 18, 19, 20, 21, 22]:
                ws.cell(row=row, column=col, value=None)
        
        for i, acct in enumerate(prior_year_data):
            row = 6 + i
            if row > 752:
                break
            ws.cell(row=row, column=16, value=int(acct['AccountNo']) if acct['AccountNo'].isdigit() else acct['AccountNo'])
            ws.cell(row=row, column=17, value=acct['Year'])
            ws.cell(row=row, column=18, value=acct['Month'])
            ws.cell(row=row, column=19, value=acct['YTD'])
            ws.cell(row=row, column=20, value=acct['MTD'])
            ws.cell(row=row, column=21, value=acct['Description'])
            ws.cell(row=row, column=22, value=acct['Type'])
        
        # --- Populate Table_TB2_ (columns Y-AD, starting row 6) ---
        for row in range(6, 752):  # Don't clear totals row
            for col in [25, 26, 27, 28, 29, 30]:
                ws.cell(row=row, column=col, value=None)
        
        for i, acct in enumerate(prior_month_data):
            row = 6 + i
            if row > 751:
                break
            ws.cell(row=row, column=25, value=int(acct['AccountNo']) if acct['AccountNo'].isdigit() else acct['AccountNo'])
            ws.cell(row=row, column=26, value='Actual')
            ws.cell(row=row, column=27, value=acct['YTD'])
            ws.cell(row=row, column=28, value=acct['MTD'])
            ws.cell(row=row, column=29, value=acct['Type'])
            ws.cell(row=row, column=30, value=acct['Description'])
        
        # --- Pre-compute all VLOOKUP formulas across all P&L sheets ---
        tb_lookup, tb1_lookup, tb2_lookup = build_lookup_dicts(
            current_data, prior_year_data, prior_month_data
        )
        precompute_all_vlookups(wb, tb_lookup, tb1_lookup, tb2_lookup)
        
        # --- Save directly with openpyxl (no XML post-processing) ---
        # This matches Amy's proven approach. openpyxl strips ODBC connections,
        # queryTables, and other problematic elements automatically.
        # The file may show a minor "repair" dialog in Excel (same as Amy's),
        # but data is fully preserved after clicking Yes.
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        month_str = f"{month:02d}"
        filename = f"{month_str}-{year}EVO.xlsx"
        
        logger.info(f"EVO export generated: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting EVO: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to export EVO file', 'message': str(e)}), 500
