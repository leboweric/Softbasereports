"""
Detailed P&L Report Route - Matches Accounting Firm Excel Format
Generates multi-tab Excel workbook with department-level GL account detail
"""

from flask import Blueprint, jsonify, request, send_file
from datetime import datetime
import logging
import calendar
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from src.config.gl_accounts_detailed import (
    DEPARTMENT_CONFIG, 
    OVERHEAD_EXPENSE_ACCOUNTS, 
    OTHER_INCOME_EXPENSE_ACCOUNTS
)
from src.routes.currie_report import get_balance_sheet_data
from flask_jwt_extended import jwt_required, get_jwt_identity
from src.utils.tenant_utils import get_tenant_db
from src.models.user import User

logger = logging.getLogger(__name__)
pl_detailed_bp = Blueprint('pl_detailed', __name__)
# sql_service is now obtained via get_tenant_db() for multi-tenant support
_sql_service = None
def get_sql_service():
    return get_tenant_db()
def get_gl_account_data(schema, account_numbers, year, month):
    """
    Get MTD and YTD data for specified GL accounts
    
    Args:
        schema: Database schema name
        account_numbers: List of GL account numbers to query
        year: Year for the report
        month: Month for the report (1-12)
    
    Returns:
        Dictionary mapping account numbers to {mtd, ytd} values
    """
    if not account_numbers:
        return {}
    
    account_list = "', '".join(account_numbers)
    
    # Query GL table for MTD and YTD values
    query = f"""
    SELECT 
        AccountNo,
        MTD,
        YTD
    FROM {schema}.GL
    WHERE Year = %s
      AND Month = %s
      AND AccountNo IN ('{account_list}')
    """
    
    try:
        results = get_sql_service().execute_query(query, [year, month])
        
        data = {}
        for row in results:
            account_no = str(row['AccountNo']).strip()
            # Revenue accounts are stored as negative (credits), so negate them
            # Expense/COS accounts are stored as positive (debits)
            mtd = float(row['MTD'] or 0)
            ytd = float(row['YTD'] or 0)
            data[account_no] = {'mtd': mtd, 'ytd': ytd}
        
        return data
    except Exception as e:
        logger.error(f"Error fetching GL account data: {e}")
        return {}


def get_department_detail(schema, dept_key, year, month):
    """
    Get detailed P&L data for a specific department
    
    Returns:
        Dictionary with sales_detail, cos_detail, totals
    """
    dept_config = DEPARTMENT_CONFIG.get(dept_key)
    if not dept_config:
        return None
    
    # Get all account numbers for this department
    sales_account_nums = [acct[0] for acct in dept_config['sales_accounts']]
    cos_account_nums = [acct[0] for acct in dept_config['cos_accounts']]
    
    # Fetch data from GL table
    all_accounts = sales_account_nums + cos_account_nums
    gl_data = get_gl_account_data(schema, all_accounts, year, month)
    
    # Build sales detail
    sales_detail = []
    total_sales_mtd = 0
    total_sales_ytd = 0
    
    for account_no, description in dept_config['sales_accounts']:
        data = gl_data.get(account_no, {'mtd': 0, 'ytd': 0})
        # Revenue is stored as negative (credit), negate to show as positive
        mtd = -data['mtd']
        ytd = -data['ytd']
        sales_detail.append({
            'account_no': account_no,
            'description': description,
            'mtd': mtd,
            'ytd': ytd
        })
        total_sales_mtd += mtd
        total_sales_ytd += ytd
    
    # Build COS detail
    cos_detail = []
    total_cos_mtd = 0
    total_cos_ytd = 0
    
    for account_no, description in dept_config['cos_accounts']:
        data = gl_data.get(account_no, {'mtd': 0, 'ytd': 0})
        # COS is stored as positive (debit)
        mtd = data['mtd']
        ytd = data['ytd']
        cos_detail.append({
            'account_no': account_no,
            'description': description,
            'mtd': mtd,
            'ytd': ytd
        })
        total_cos_mtd += mtd
        total_cos_ytd += ytd
    
    # Calculate gross profit
    gross_profit_mtd = total_sales_mtd - total_cos_mtd
    gross_profit_ytd = total_sales_ytd - total_cos_ytd
    
    return {
        'dept_code': dept_config['dept_code'],
        'dept_name': dept_config['dept_name'],
        'tab_name': dept_config['tab_name'],
        'sales_detail': sales_detail,
        'cos_detail': cos_detail,
        'total_sales_mtd': total_sales_mtd,
        'total_sales_ytd': total_sales_ytd,
        'total_cos_mtd': total_cos_mtd,
        'total_cos_ytd': total_cos_ytd,
        'gross_profit_mtd': gross_profit_mtd,
        'gross_profit_ytd': gross_profit_ytd,
        'gross_margin_mtd': (gross_profit_mtd / total_sales_mtd * 100) if total_sales_mtd else 0,
        'gross_margin_ytd': (gross_profit_ytd / total_sales_ytd * 100) if total_sales_ytd else 0
    }


def get_overhead_expenses(schema, year, month):
    """Get overhead expense data organized by category"""
    all_expense_accounts = []
    for category_accounts in OVERHEAD_EXPENSE_ACCOUNTS.values():
        all_expense_accounts.extend([acct[0] for acct in category_accounts])
    
    gl_data = get_gl_account_data(schema, all_expense_accounts, year, month)
    
    expense_data = {}
    total_mtd = 0
    total_ytd = 0
    
    for category, accounts in OVERHEAD_EXPENSE_ACCOUNTS.items():
        category_detail = []
        category_mtd = 0
        category_ytd = 0
        
        for account_no, description in accounts:
            data = gl_data.get(account_no, {'mtd': 0, 'ytd': 0})
            mtd = data['mtd']
            ytd = data['ytd']
            category_detail.append({
                'account_no': account_no,
                'description': description,
                'mtd': mtd,
                'ytd': ytd
            })
            category_mtd += mtd
            category_ytd += ytd
        
        expense_data[category] = {
            'detail': category_detail,
            'total_mtd': category_mtd,
            'total_ytd': category_ytd
        }
        total_mtd += category_mtd
        total_ytd += category_ytd
    
    expense_data['total_overhead_mtd'] = total_mtd
    expense_data['total_overhead_ytd'] = total_ytd
    
    return expense_data


def get_other_income_expense(schema, year, month):
    """Get other income and expense data"""
    all_accounts = []
    for category_accounts in OTHER_INCOME_EXPENSE_ACCOUNTS.values():
        all_accounts.extend([acct[0] for acct in category_accounts])
    
    gl_data = get_gl_account_data(schema, all_accounts, year, month)
    
    result = {}
    
    for category, accounts in OTHER_INCOME_EXPENSE_ACCOUNTS.items():
        category_detail = []
        category_mtd = 0
        category_ytd = 0
        
        for account_no, description in accounts:
            data = gl_data.get(account_no, {'mtd': 0, 'ytd': 0})
            mtd = data['mtd']
            ytd = data['ytd']
            category_detail.append({
                'account_no': account_no,
                'description': description,
                'mtd': mtd,
                'ytd': ytd
            })
            category_mtd += mtd
            category_ytd += ytd
        
        result[category] = {
            'detail': category_detail,
            'total_mtd': category_mtd,
            'total_ytd': category_ytd
        }
    
    # Calculate total other income & expense
    result['total_mtd'] = (result.get('other_income', {}).get('total_mtd', 0) + 
                          result.get('other_expense', {}).get('total_mtd', 0))
    result['total_ytd'] = (result.get('other_income', {}).get('total_ytd', 0) + 
                          result.get('other_expense', {}).get('total_ytd', 0))
    
    return result


def create_department_worksheet(wb, dept_data, year, month):
    """Create a worksheet for a department P&L"""
    ws = wb.create_sheet(title=dept_data['tab_name'])
    
    # Define styles
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    money_format = '#,##0.00'
    percent_format = '0.00%'
    
    thin_border = Border(
        bottom=Side(style='thin')
    )
    
    # Row 1: Title
    ws['A1'] = f"Income Statement - Departmental"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    # Row 2: Month/Year
    month_name = calendar.month_name[month]
    ws['A2'] = f"{month_name} {year}"
    ws['A2'].font = header_font
    
    # Row 3: Department Code
    ws['A3'] = dept_data['dept_code']
    ws['A3'].font = header_font
    
    # Row 4: Headers
    ws['A4'] = dept_data['dept_name']
    ws['A4'].font = header_font
    ws['C4'] = 'MTD'
    ws['C4'].font = header_font
    ws['D4'] = '%'
    ws['D4'].font = header_font
    ws['E4'] = 'YTD'
    ws['E4'].font = header_font
    ws['F4'] = '%'
    ws['F4'].font = header_font
    
    # Row 5: Sales header
    ws['A5'] = 'Sales'
    ws['A5'].font = header_font
    
    current_row = 7
    
    # Sales detail rows
    for item in dept_data['sales_detail']:
        ws[f'A{current_row}'] = item['account_no']
        ws[f'B{current_row}'] = item['description']
        ws[f'C{current_row}'] = item['mtd']
        ws[f'C{current_row}'].number_format = money_format
        if dept_data['total_sales_mtd'] != 0:
            ws[f'D{current_row}'] = item['mtd'] / dept_data['total_sales_mtd']
        else:
            ws[f'D{current_row}'] = 0
        ws[f'D{current_row}'].number_format = percent_format
        ws[f'E{current_row}'] = item['ytd']
        ws[f'E{current_row}'].number_format = money_format
        if dept_data['total_sales_ytd'] != 0:
            ws[f'F{current_row}'] = item['ytd'] / dept_data['total_sales_ytd']
        else:
            ws[f'F{current_row}'] = 0
        ws[f'F{current_row}'].number_format = percent_format
        current_row += 1
    
    # Total Sales row
    current_row += 1
    total_sales_row = current_row
    ws[f'A{current_row}'] = 'Total Sales'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = dept_data['total_sales_mtd']
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    ws[f'D{current_row}'] = 1.0
    ws[f'D{current_row}'].number_format = percent_format
    ws[f'E{current_row}'] = dept_data['total_sales_ytd']
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    ws[f'F{current_row}'] = 1.0
    ws[f'F{current_row}'].number_format = percent_format
    current_row += 2
    
    # Cost of Sales header
    ws[f'A{current_row}'] = 'Cost of Sales'
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    # COS detail rows
    for item in dept_data['cos_detail']:
        ws[f'A{current_row}'] = item['account_no']
        ws[f'B{current_row}'] = item['description']
        ws[f'C{current_row}'] = item['mtd']
        ws[f'C{current_row}'].number_format = money_format
        if dept_data['total_sales_mtd'] != 0:
            ws[f'D{current_row}'] = item['mtd'] / dept_data['total_sales_mtd']
        else:
            ws[f'D{current_row}'] = 0
        ws[f'D{current_row}'].number_format = percent_format
        ws[f'E{current_row}'] = item['ytd']
        ws[f'E{current_row}'].number_format = money_format
        if dept_data['total_sales_ytd'] != 0:
            ws[f'F{current_row}'] = item['ytd'] / dept_data['total_sales_ytd']
        else:
            ws[f'F{current_row}'] = 0
        ws[f'F{current_row}'].number_format = percent_format
        current_row += 1
    
    # Total COGS row
    current_row += 1
    ws[f'A{current_row}'] = 'Total Cost of Goods Sold'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = dept_data['total_cos_mtd']
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    if dept_data['total_sales_mtd'] != 0:
        ws[f'D{current_row}'] = dept_data['total_cos_mtd'] / dept_data['total_sales_mtd']
    else:
        ws[f'D{current_row}'] = 0
    ws[f'D{current_row}'].number_format = percent_format
    ws[f'E{current_row}'] = dept_data['total_cos_ytd']
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    if dept_data['total_sales_ytd'] != 0:
        ws[f'F{current_row}'] = dept_data['total_cos_ytd'] / dept_data['total_sales_ytd']
    else:
        ws[f'F{current_row}'] = 0
    ws[f'F{current_row}'].number_format = percent_format
    current_row += 2
    
    # Gross Profit row
    ws[f'A{current_row}'] = 'Gross Profit'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = dept_data['gross_profit_mtd']
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    ws[f'D{current_row}'] = dept_data['gross_margin_mtd'] / 100
    ws[f'D{current_row}'].number_format = percent_format
    ws[f'E{current_row}'] = dept_data['gross_profit_ytd']
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    ws[f'F{current_row}'] = dept_data['gross_margin_ytd'] / 100
    ws[f'F{current_row}'].number_format = percent_format
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    
    return ws


def create_inhouse_worksheet(wb, dept_data, expense_data, other_data, year, month, inhouse_dept_data=None):
    """Create the In House / Administrative worksheet with expenses
    
    Args:
        wb: Workbook object
        dept_data: Dictionary with total_gross_profit_mtd and total_gross_profit_ytd
        expense_data: Overhead expense data
        other_data: Other income/expense data
        year: Report year
        month: Report month
        inhouse_dept_data: Optional In House department Sales/COGS data
    """
    ws = wb.create_sheet(title='P&L In House')
    
    # Define styles
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    money_format = '#,##0.00'
    percent_format = '0.00%'
    
    # Row 1: Title
    ws['A1'] = f"Income Statement - Departmental"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    # Row 2: Month/Year
    month_name = calendar.month_name[month]
    ws['A2'] = f"{month_name} {year}"
    ws['A2'].font = header_font
    
    # Row 3: Department Code
    ws['A3'] = 90
    ws['A3'].font = header_font
    
    # Row 4: Headers
    ws['A4'] = 'In House / Administrative'
    ws['A4'].font = header_font
    ws['C4'] = 'MTD'
    ws['C4'].font = header_font
    ws['E4'] = 'YTD'
    ws['E4'].font = header_font
    
    current_row = 5
    
    # === SALES SECTION (to match accounting firm format) ===
    ws[f'A{current_row}'] = 'Sales'
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    # Sales line items (if available)
    inhouse_sales_mtd = 0
    inhouse_sales_ytd = 0
    if inhouse_dept_data and 'sales_detail' in inhouse_dept_data:
        for item in inhouse_dept_data['sales_detail']:
            if item['mtd'] != 0 or item['ytd'] != 0:
                ws[f'A{current_row}'] = item['account_no']
                ws[f'B{current_row}'] = item['description']
                ws[f'C{current_row}'] = item['mtd']
                ws[f'C{current_row}'].number_format = money_format
                ws[f'E{current_row}'] = item['ytd']
                ws[f'E{current_row}'].number_format = money_format
                current_row += 1
        inhouse_sales_mtd = inhouse_dept_data.get('total_sales_mtd', 0)
        inhouse_sales_ytd = inhouse_dept_data.get('total_sales_ytd', 0)
    
    current_row += 1
    ws[f'A{current_row}'] = 'Total Sales'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = inhouse_sales_mtd
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    ws[f'E{current_row}'] = inhouse_sales_ytd
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    current_row += 2
    
    # === COST OF SALES SECTION ===
    ws[f'A{current_row}'] = 'Cost of Sales'
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    # COS line items (if available)
    inhouse_cos_mtd = 0
    inhouse_cos_ytd = 0
    if inhouse_dept_data and 'cos_detail' in inhouse_dept_data:
        for item in inhouse_dept_data['cos_detail']:
            if item['mtd'] != 0 or item['ytd'] != 0:
                ws[f'A{current_row}'] = item['account_no']
                ws[f'B{current_row}'] = item['description']
                ws[f'C{current_row}'] = item['mtd']
                ws[f'C{current_row}'].number_format = money_format
                ws[f'E{current_row}'] = item['ytd']
                ws[f'E{current_row}'].number_format = money_format
                current_row += 1
        inhouse_cos_mtd = inhouse_dept_data.get('total_cos_mtd', 0)
        inhouse_cos_ytd = inhouse_dept_data.get('total_cos_ytd', 0)
    
    current_row += 1
    ws[f'A{current_row}'] = 'Total Cost of Goods Sold'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = inhouse_cos_mtd
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    ws[f'E{current_row}'] = inhouse_cos_ytd
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    current_row += 2
    
    # === GROSS PROFIT (In House department only) ===
    inhouse_gp_mtd = inhouse_sales_mtd - inhouse_cos_mtd
    inhouse_gp_ytd = inhouse_sales_ytd - inhouse_cos_ytd
    ws[f'A{current_row}'] = 'Gross Profit'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = inhouse_gp_mtd
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    ws[f'E{current_row}'] = inhouse_gp_ytd
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    current_row += 2
    
    # === OVERHEAD / SG&A SECTION ===
    ws[f'A{current_row}'] = 'Overhead / Sales, General & Administrative'
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    # Expense categories - MUST match keys in OVERHEAD_EXPENSE_ACCOUNTS
    category_names = {
        'depreciation': 'Depreciation',
        'salaries_wages': 'Salaries & Wages',
        'payroll_benefits': 'Payroll Benefits',
        'rent_facilities': 'Rent & Facilities',
        'utilities': 'Utilities',
        'insurance': 'Insurance',
        'marketing': 'Marketing & Advertising',
        'professional_fees': 'Professional Fees',
        'office_admin': 'Office & Administrative',
        'vehicle_equipment': 'Vehicle & Equipment',
        'other_expenses': 'Other Expenses',
    }
    
    for category, display_name in category_names.items():
        if category in expense_data:
            cat_data = expense_data[category]
            
            # Category detail rows
            for item in cat_data['detail']:
                ws[f'A{current_row}'] = item['account_no']
                ws[f'B{current_row}'] = item['description']
                ws[f'C{current_row}'] = item['mtd']
                ws[f'C{current_row}'].number_format = money_format
                ws[f'E{current_row}'] = item['ytd']
                ws[f'E{current_row}'].number_format = money_format
                current_row += 1
            
            # Category subtotal
            ws[f'B{current_row}'] = f'Total {display_name}'
            ws[f'B{current_row}'].font = header_font
            ws[f'C{current_row}'] = cat_data['total_mtd']
            ws[f'C{current_row}'].number_format = money_format
            ws[f'C{current_row}'].font = header_font
            ws[f'E{current_row}'] = cat_data['total_ytd']
            ws[f'E{current_row}'].number_format = money_format
            ws[f'E{current_row}'].font = header_font
            current_row += 1
    
    current_row += 1
    
    # Total Overhead Expenses
    ws[f'A{current_row}'] = 'Total Overhead Expenses'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = expense_data.get('total_overhead_mtd', 0)
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    ws[f'E{current_row}'] = expense_data.get('total_overhead_ytd', 0)
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    current_row += 2
    
    # Operating Profit (In House Gross Profit minus Overhead Expenses)
    operating_profit_mtd = inhouse_gp_mtd - expense_data.get('total_overhead_mtd', 0)
    operating_profit_ytd = inhouse_gp_ytd - expense_data.get('total_overhead_ytd', 0)
    
    ws[f'A{current_row}'] = 'Operating Profit'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = operating_profit_mtd
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    ws[f'E{current_row}'] = operating_profit_ytd
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    current_row += 2
    
    # Other Income & Expense header
    ws[f'A{current_row}'] = 'Other Income & Expense'
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    # Other Income
    if 'other_income' in other_data:
        for item in other_data['other_income']['detail']:
            ws[f'A{current_row}'] = item['account_no']
            ws[f'B{current_row}'] = item['description']
            ws[f'C{current_row}'] = item['mtd']
            ws[f'C{current_row}'].number_format = money_format
            ws[f'E{current_row}'] = item['ytd']
            ws[f'E{current_row}'].number_format = money_format
            current_row += 1
        
        ws[f'C{current_row}'] = other_data['other_income']['total_mtd']
        ws[f'C{current_row}'].number_format = money_format
        ws[f'C{current_row}'].font = header_font
        ws[f'E{current_row}'] = other_data['other_income']['total_ytd']
        ws[f'E{current_row}'].number_format = money_format
        ws[f'E{current_row}'].font = header_font
        current_row += 2
    
    # Other Expense
    if 'other_expense' in other_data:
        for item in other_data['other_expense']['detail']:
            ws[f'A{current_row}'] = item['account_no']
            ws[f'B{current_row}'] = item['description']
            ws[f'C{current_row}'] = item['mtd']
            ws[f'C{current_row}'].number_format = money_format
            ws[f'E{current_row}'] = item['ytd']
            ws[f'E{current_row}'].number_format = money_format
            current_row += 1
        
        ws[f'C{current_row}'] = other_data['other_expense']['total_mtd']
        ws[f'C{current_row}'].number_format = money_format
        ws[f'C{current_row}'].font = header_font
        ws[f'E{current_row}'] = other_data['other_expense']['total_ytd']
        ws[f'E{current_row}'].number_format = money_format
        ws[f'E{current_row}'].font = header_font
        current_row += 2
    
    # Total Other Income & Expense
    ws[f'A{current_row}'] = 'Total Other Income & Expense'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = other_data.get('total_mtd', 0)
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    ws[f'E{current_row}'] = other_data.get('total_ytd', 0)
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    current_row += 2
    
    # Net Profit
    net_profit_mtd = operating_profit_mtd - other_data.get('total_mtd', 0)
    net_profit_ytd = operating_profit_ytd - other_data.get('total_ytd', 0)
    
    ws[f'A{current_row}'] = 'Net Profit'
    ws[f'A{current_row}'].font = header_font
    ws[f'C{current_row}'] = net_profit_mtd
    ws[f'C{current_row}'].number_format = money_format
    ws[f'C{current_row}'].font = header_font
    ws[f'E{current_row}'] = net_profit_ytd
    ws[f'E{current_row}'].number_format = money_format
    ws[f'E{current_row}'].font = header_font
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    
    return ws


def create_consolidated_worksheet(wb, all_dept_data, expense_data, other_data, year, month):
    """Create the consolidated P&L summary worksheet"""
    ws = wb.create_sheet(title='Profit & Loss Consolidated')
    
    # Define styles
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    money_format = '#,##0.00'
    percent_format = '0.00%'
    
    # Row 1: Title
    ws['A1'] = 'Profit & Loss Consolidated'
    ws['A1'].font = title_font
    
    # Row 2: Month/Year
    month_name = calendar.month_name[month]
    ws['A2'] = f"For the Period Ended {month_name} {year}"
    
    # Row 4: MTD Summary header
    ws['B4'] = 'MTD Summary'
    ws['B4'].font = header_font
    
    # Row 5: Column headers
    headers = ['', 'New Equipment', 'Used Equipment', 'Parts', 'Service', 'Rental', 'Transportation', 'In House', 'Total']
    dept_codes = ['', '10', '20', '30', '40', '60', '80', '90', '']
    
    for col, (header, code) in enumerate(zip(headers, dept_codes), 2):
        ws.cell(row=4, column=col, value=code).font = header_font
        ws.cell(row=5, column=col, value=header).font = header_font
    
    # Data rows for MTD
    row_labels = ['Income', 'Cost of Goods Sold', 'Gross Profit', 'Gross Margin', 
                  'Overhead Expenses', 'Operating Profit', 'Operating Margin',
                  'Other Income & Expense', 'Net Profit', 'Net Margin']
    
    dept_keys = ['new_equipment', 'used_equipment', 'parts', 'service', 'rental', 'transportation', 'in_house']
    
    current_row = 6
    
    # Calculate totals
    # Note: Other income (7xxxxx accounts) includes contra-revenue items like A/R Discounts
    # which reduce total revenue. The dashboard adds other_income to revenue (negative value reduces it)
    total_other_mtd = other_data.get('total_mtd', 0)
    total_income_mtd = sum(all_dept_data[k]['total_sales_mtd'] for k in dept_keys if k in all_dept_data) - total_other_mtd
    total_cogs_mtd = sum(all_dept_data[k]['total_cos_mtd'] for k in dept_keys if k in all_dept_data)
    total_gross_profit_mtd = total_income_mtd - total_cogs_mtd
    total_overhead_mtd = expense_data.get('total_overhead_mtd', 0)
    total_operating_profit_mtd = total_gross_profit_mtd - total_overhead_mtd
    total_net_profit_mtd = total_operating_profit_mtd
    
    # Income row
    ws.cell(row=current_row, column=2, value='Income')
    for col, key in enumerate(dept_keys, 3):
        if key in all_dept_data:
            ws.cell(row=current_row, column=col, value=all_dept_data[key]['total_sales_mtd']).number_format = money_format
    ws.cell(row=current_row, column=10, value=total_income_mtd).number_format = money_format
    current_row += 1
    
    # COGS row
    ws.cell(row=current_row, column=2, value='Cost of Goods Sold')
    for col, key in enumerate(dept_keys, 3):
        if key in all_dept_data:
            ws.cell(row=current_row, column=col, value=all_dept_data[key]['total_cos_mtd']).number_format = money_format
    ws.cell(row=current_row, column=10, value=total_cogs_mtd).number_format = money_format
    current_row += 1
    
    # Gross Profit row
    ws.cell(row=current_row, column=2, value='Gross Profit')
    for col, key in enumerate(dept_keys, 3):
        if key in all_dept_data:
            ws.cell(row=current_row, column=col, value=all_dept_data[key]['gross_profit_mtd']).number_format = money_format
    ws.cell(row=current_row, column=10, value=total_gross_profit_mtd).number_format = money_format
    current_row += 1
    
    # Gross Margin row
    ws.cell(row=current_row, column=2, value='Gross Margin').font = header_font
    for col, key in enumerate(dept_keys, 3):
        if key in all_dept_data and all_dept_data[key]['total_sales_mtd'] != 0:
            margin = all_dept_data[key]['gross_profit_mtd'] / all_dept_data[key]['total_sales_mtd']
            ws.cell(row=current_row, column=col, value=margin).number_format = percent_format
    if total_income_mtd != 0:
        ws.cell(row=current_row, column=10, value=total_gross_profit_mtd / total_income_mtd).number_format = percent_format
    current_row += 1
    
    # Overhead row
    ws.cell(row=current_row, column=2, value='Overhead Expenses')
    ws.cell(row=current_row, column=9, value=total_overhead_mtd).number_format = money_format
    ws.cell(row=current_row, column=10, value=total_overhead_mtd).number_format = money_format
    current_row += 1
    
    # Operating Profit row
    ws.cell(row=current_row, column=2, value='Operating Profit')
    ws.cell(row=current_row, column=10, value=total_operating_profit_mtd).number_format = money_format
    current_row += 1
    
    # Operating Margin row
    ws.cell(row=current_row, column=2, value='Operating Margin').font = header_font
    if total_income_mtd != 0:
        ws.cell(row=current_row, column=10, value=total_operating_profit_mtd / total_income_mtd).number_format = percent_format
    current_row += 1
    
    # Other Income & Expense row
    ws.cell(row=current_row, column=2, value='Other Income & Expense')
    ws.cell(row=current_row, column=9, value=total_other_mtd).number_format = money_format
    ws.cell(row=current_row, column=10, value=total_other_mtd).number_format = money_format
    current_row += 1
    
    # Net Profit row
    ws.cell(row=current_row, column=2, value='Net Profit').font = header_font
    ws.cell(row=current_row, column=10, value=total_net_profit_mtd).number_format = money_format
    ws.cell(row=current_row, column=10).font = header_font
    current_row += 1
    
    # Net Margin row
    ws.cell(row=current_row, column=2, value='Net Margin').font = header_font
    if total_income_mtd != 0:
        ws.cell(row=current_row, column=10, value=total_net_profit_mtd / total_income_mtd).number_format = percent_format
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    return ws


def create_balance_sheet_worksheet(wb, year, month):
    """Create the Balance Sheet worksheet matching accounting firm's format"""
    ws = wb.create_sheet(title='Balance Sheet')
    
    # Define styles
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    money_format = '#,##0.00'
    
    # Get balance sheet data
    as_of_date = f"{year}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"
    bs_data = get_balance_sheet_data(as_of_date)
    
    # Helper function to sum account balances
    def sum_accounts(account_list):
        return sum(acc.get('balance', 0) for acc in account_list)
    
    # Row 1: Title
    month_name = calendar.month_name[month]
    ws['A1'] = f"Balance Sheet\n As of {month_name} {year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 35
    
    current_row = 4
    
    # ============= ASSETS =============
    ws.cell(row=current_row, column=1, value='Assets').font = section_font
    ws.cell(row=current_row, column=3, value='Total').font = header_font
    current_row += 2
    
    assets = bs_data['assets']
    
    # Cash & Equivalents
    ws.cell(row=current_row, column=2, value='Cash & Equivalents').font = header_font
    current_row += 2
    
    cash_accounts = assets['current_assets']['cash']
    cash_total = 0
    for acc in sorted(cash_accounts, key=lambda x: x['account']):
        ws.cell(row=current_row, column=2, value=acc['description'])
        ws.cell(row=current_row, column=3, value=acc['balance']).number_format = money_format
        cash_total += acc['balance']
        current_row += 1
    
    ws.cell(row=current_row, column=2, value=' Total')
    ws.cell(row=current_row, column=3, value=cash_total).number_format = money_format
    current_row += 2
    
    ws.cell(row=current_row, column=2, value='Total Cash & Equivalents').font = header_font
    ws.cell(row=current_row, column=3, value=cash_total).number_format = money_format
    ws.cell(row=current_row, column=3).font = header_font
    current_row += 2
    
    # Current Assets
    ws.cell(row=current_row, column=2, value='Current Assets').font = header_font
    current_row += 1
    
    # Accounts Receivable - exclude PARTS RETURN (goes to WIP)
    ws.cell(row=current_row, column=2, value='Accounts Receivable').font = header_font
    current_row += 1
    
    ar_accounts = assets['current_assets']['accounts_receivable']
    ar_total = 0
    wip_from_ar = []  # Accounts that should go to WIP instead
    for acc in sorted(ar_accounts, key=lambda x: x['account']):
        desc = acc['description'].upper()
        if 'RETURN' in desc and 'PROCESS' in desc:
            wip_from_ar.append(acc)  # Move to WIP section
        else:
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=acc['balance']).number_format = money_format
            ar_total += acc['balance']
            current_row += 1
    
    ws.cell(row=current_row, column=2, value='Accounts Receivable Total').font = header_font
    ws.cell(row=current_row, column=3, value=ar_total).number_format = money_format
    current_row += 2
    
    # Inventory - exclude WORK-IN-PROCESS (goes to WIP section)
    ws.cell(row=current_row, column=2, value='Inventory').font = header_font
    current_row += 1
    
    inventory_accounts = assets['current_assets']['inventory']
    inventory_total = 0
    wip_from_inventory = []  # Accounts that should go to WIP instead
    for acc in sorted(inventory_accounts, key=lambda x: x['account']):
        desc = acc['description'].upper()
        if 'WORK' in desc and 'PROCESS' in desc:
            wip_from_inventory.append(acc)  # Move to WIP section
        else:
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=acc['balance']).number_format = money_format
            inventory_total += acc['balance']
            current_row += 1
    
    ws.cell(row=current_row, column=2, value='Inventory Total').font = header_font
    ws.cell(row=current_row, column=3, value=inventory_total).number_format = money_format
    current_row += 2
    
    # Other Current Assets (Prepaid, WIP, Deposits)
    other_current = assets['current_assets']['other_current']
    other_current_total = 0
    
    # Group by type
    prepaid_accounts = []
    wip_accounts = []
    deposit_accounts = []
    other_accounts = []
    
    for acc in other_current:
        desc = acc['description'].upper()
        if 'PREPAID' in desc:
            prepaid_accounts.append(acc)
        elif 'WORK' in desc or 'WIP' in desc or 'PROCESS' in desc or 'RETURN' in desc:
            wip_accounts.append(acc)
        elif 'DEPOSIT' in desc:
            deposit_accounts.append(acc)
        else:
            other_accounts.append(acc)
    
    # Prepaid Expense
    if prepaid_accounts:
        ws.cell(row=current_row, column=2, value='Prepaid Expense').font = header_font
        current_row += 1
        prepaid_total = 0
        for acc in sorted(prepaid_accounts, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=acc['balance']).number_format = money_format
            prepaid_total += acc['balance']
            current_row += 1
        ws.cell(row=current_row, column=2, value='Prepaid Expense Total').font = header_font
        ws.cell(row=current_row, column=3, value=prepaid_total).number_format = money_format
        other_current_total += prepaid_total
        current_row += 2
    
    # Work In Process - include accounts moved from AR and Inventory
    all_wip_accounts = wip_accounts + wip_from_ar + wip_from_inventory
    if all_wip_accounts:
        ws.cell(row=current_row, column=2, value='Work In Process').font = header_font
        current_row += 1
        wip_total = 0
        for acc in sorted(all_wip_accounts, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=acc['balance']).number_format = money_format
            wip_total += acc['balance']
            current_row += 1
        ws.cell(row=current_row, column=2, value='Work In Process Total').font = header_font
        ws.cell(row=current_row, column=3, value=wip_total).number_format = money_format
        other_current_total += wip_total
        current_row += 2
    
    # Deposits
    if deposit_accounts:
        ws.cell(row=current_row, column=2, value='Deposits').font = header_font
        current_row += 1
        for acc in sorted(deposit_accounts, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=acc['balance']).number_format = money_format
            other_current_total += acc['balance']
            current_row += 1
        current_row += 1
    
    # Other current assets
    for acc in sorted(other_accounts, key=lambda x: x['account']):
        ws.cell(row=current_row, column=2, value=acc['description'])
        ws.cell(row=current_row, column=3, value=acc['balance']).number_format = money_format
        other_current_total += acc['balance']
        current_row += 1
    
    # Total Current Assets (excludes Cash & Equivalents per accounting firm format)
    total_current_assets = ar_total + inventory_total + other_current_total
    ws.cell(row=current_row, column=2, value='Total Current Assets').font = header_font
    ws.cell(row=current_row, column=3, value=total_current_assets).number_format = money_format
    ws.cell(row=current_row, column=3).font = header_font
    current_row += 2
    
    # Property and Equipment
    ws.cell(row=current_row, column=2, value='Property and Equipment').font = header_font
    current_row += 1
    
    fixed_assets = assets['fixed_assets']
    fixed_total_gross = 0
    fixed_total_deprec = 0
    
    # Group fixed assets by category
    furniture_accounts = []
    leasehold_accounts = []
    machinery_accounts = []
    rental_accounts = []
    vehicle_accounts = []
    other_fixed_accounts = []
    
    rou_lease_accounts = []  # RIGHT OF USE accounts go to Other Assets
    for acc in fixed_assets:
        desc = acc['description'].upper()
        if 'RIGHT OF USE' in desc or 'ROU' in desc:
            rou_lease_accounts.append(acc)  # Move to Other Assets
        elif 'FURNITURE' in desc or 'FIXTURE' in desc:
            furniture_accounts.append(acc)
        elif 'LEASEHOLD' in desc:
            leasehold_accounts.append(acc)
        elif 'MACHINERY' in desc or ('EQUIPMENT' in desc and 'RENTAL' not in desc):
            machinery_accounts.append(acc)
        elif 'RENTAL' in desc:
            rental_accounts.append(acc)
        elif 'VEHICLE' in desc:
            vehicle_accounts.append(acc)
        else:
            other_fixed_accounts.append(acc)
    
    def write_fixed_asset_group(accounts, group_name):
        nonlocal current_row, fixed_total_gross, fixed_total_deprec
        if not accounts:
            return
        ws.cell(row=current_row, column=2, value=group_name).font = header_font
        current_row += 1
        group_total = 0
        for acc in sorted(accounts, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=acc['balance']).number_format = money_format
            group_total += acc['balance']
            if 'DEPREC' in acc['description'].upper() or 'ACCUM' in acc['description'].upper():
                fixed_total_deprec += acc['balance']
            else:
                fixed_total_gross += acc['balance']
            current_row += 1
        ws.cell(row=current_row, column=2, value=f'{group_name} Total').font = header_font
        ws.cell(row=current_row, column=3, value=group_total).number_format = money_format
        current_row += 1
    
    write_fixed_asset_group(furniture_accounts, 'Furniture & Fixtures')
    write_fixed_asset_group(leasehold_accounts, 'Leasehold Improvements')
    write_fixed_asset_group(machinery_accounts, 'Machinery & Equipment')
    write_fixed_asset_group(rental_accounts, 'Rental Equipment')
    write_fixed_asset_group(vehicle_accounts, 'Vehicles')
    write_fixed_asset_group(other_fixed_accounts, 'Other Fixed Assets')
    
    current_row += 1
    total_fixed = fixed_total_gross + fixed_total_deprec
    ws.cell(row=current_row, column=2, value='Total Property and Equipment').font = header_font
    ws.cell(row=current_row, column=3, value=fixed_total_gross).number_format = money_format
    current_row += 1
    ws.cell(row=current_row, column=2, value='Less Depreciation')
    ws.cell(row=current_row, column=3, value=fixed_total_deprec).number_format = money_format
    current_row += 1
    ws.cell(row=current_row, column=2, value='Property and Equipment, net').font = header_font
    ws.cell(row=current_row, column=3, value=total_fixed).number_format = money_format
    ws.cell(row=current_row, column=3).font = header_font
    current_row += 2
    
    # Other Assets - include ROU lease accounts moved from Fixed Assets
    other_assets = assets['other_assets']
    all_other_assets = list(other_assets) + rou_lease_accounts
    if all_other_assets:
        ws.cell(row=current_row, column=2, value='Other Assets').font = header_font
        current_row += 1
        ws.cell(row=current_row, column=2, value='Other Assets').font = header_font
        current_row += 1
        other_assets_total = 0
        for acc in sorted(all_other_assets, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=acc['balance']).number_format = money_format
            other_assets_total += acc['balance']
            current_row += 1
        ws.cell(row=current_row, column=2, value='Other Assets Total').font = header_font
        ws.cell(row=current_row, column=3, value=other_assets_total).number_format = money_format
        current_row += 2
        ws.cell(row=current_row, column=2, value='Total Other Assets').font = header_font
        ws.cell(row=current_row, column=3, value=other_assets_total).number_format = money_format
        current_row += 2
    else:
        other_assets_total = 0
    
    # Total Assets (Cash + Current Assets + Fixed Assets + Other Assets)
    total_assets = cash_total + total_current_assets + total_fixed + other_assets_total
    ws.cell(row=current_row, column=2, value='Total Assets').font = section_font
    ws.cell(row=current_row, column=3, value=total_assets).number_format = money_format
    ws.cell(row=current_row, column=3).font = section_font
    current_row += 2
    
    # ============= LIABILITIES AND STOCKHOLDERS' EQUITY =============
    ws.cell(row=current_row, column=1, value="Liabilities and Stockholders' Equity").font = section_font
    ws.cell(row=current_row, column=3, value='Total').font = header_font
    current_row += 2
    
    liabilities = bs_data['liabilities']
    
    # Current Liabilities
    ws.cell(row=current_row, column=2, value='Current Liabilities').font = header_font
    current_row += 1
    
    current_liab = liabilities['current_liabilities']
    long_term_liab = list(liabilities['long_term_liabilities'])  # Start with existing long-term
    
    # Group current liabilities - move certain accounts to long-term
    ap_accounts = []
    accrued_accounts = []
    payroll_accounts = []
    sales_tax_accounts = []
    credit_card_accounts = []
    other_current_liab = []
    moved_to_lt = []  # Accounts that should be long-term
    
    for acc in current_liab:
        desc = acc['description'].upper()
        # These accounts should be in Long-term Liabilities per accounting firm
        if 'FLOOR PLAN' in desc:
            moved_to_lt.append(acc)
        elif 'OPERATING LEASE' in desc and 'LT' in desc:
            moved_to_lt.append(acc)
        elif 'NP - EXEC' in desc or 'NP - SCALE' in desc:
            moved_to_lt.append(acc)
        elif 'ACCOUNTS PAYABLE' in desc or 'A/P' in desc:
            ap_accounts.append(acc)
        elif 'ACCRUED' in desc or 'PAYROLL' in desc or 'SALARY' in desc or 'WAGES' in desc or 'DEFERRED COMP' in desc or 'GARNISH' in desc:
            payroll_accounts.append(acc)
        elif 'SALES TAX' in desc or 'PAR INCOME TAX' in desc:
            sales_tax_accounts.append(acc)
        elif 'CREDIT CARD' in desc:
            credit_card_accounts.append(acc)
        elif 'TRUCKS PURCHASED' in desc:
            accrued_accounts.append(acc)
        else:
            other_current_liab.append(acc)
    
    # Add moved accounts to long-term liabilities
    long_term_liab.extend(moved_to_lt)
    
    current_liab_total = 0
    
    # Accounts Payable (negate balance - liabilities stored as negative credits)
    if ap_accounts:
        ws.cell(row=current_row, column=2, value='Accounts Payable').font = header_font
        current_row += 1
        for acc in sorted(ap_accounts, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
            current_liab_total += -acc['balance']
            current_row += 1
        current_row += 1
    
    # Accrued Expenses (negate balance)
    if accrued_accounts or other_current_liab:
        ws.cell(row=current_row, column=2, value='Accrued Expenses').font = header_font
        current_row += 1
        for acc in sorted(accrued_accounts + other_current_liab, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
            current_liab_total += -acc['balance']
            current_row += 1
        current_row += 1
    
    # Accrued Payroll (negate balance)
    if payroll_accounts:
        ws.cell(row=current_row, column=2, value='Accrued Payroll').font = header_font
        current_row += 1
        payroll_total = 0
        for acc in sorted(payroll_accounts, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
            payroll_total += -acc['balance']
            current_row += 1
        ws.cell(row=current_row, column=2, value='Accrued Payroll Total').font = header_font
        ws.cell(row=current_row, column=3, value=payroll_total).number_format = money_format
        current_liab_total += payroll_total
        current_row += 2
    
    # Sales Tax Payable (negate balance)
    if sales_tax_accounts:
        ws.cell(row=current_row, column=2, value='Sales Tax Payable').font = header_font
        current_row += 1
        sales_tax_total = 0
        for acc in sorted(sales_tax_accounts, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
            sales_tax_total += -acc['balance']
            current_row += 1
        ws.cell(row=current_row, column=2, value='Sales Tax Payable Total').font = header_font
        ws.cell(row=current_row, column=3, value=sales_tax_total).number_format = money_format
        current_liab_total += sales_tax_total
        current_row += 2
    
    # Credit Cards (negate balance)
    if credit_card_accounts:
        ws.cell(row=current_row, column=2, value='Credit Cards').font = header_font
        current_row += 1
        cc_total = 0
        for acc in sorted(credit_card_accounts, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
            cc_total += -acc['balance']
            current_row += 1
        ws.cell(row=current_row, column=2, value='Credit Cards Total').font = header_font
        ws.cell(row=current_row, column=3, value=cc_total).number_format = money_format
        current_liab_total += cc_total
        current_row += 2
    
    ws.cell(row=current_row, column=2, value='Total Current Liabilities').font = header_font
    ws.cell(row=current_row, column=3, value=current_liab_total).number_format = money_format
    ws.cell(row=current_row, column=3).font = header_font
    current_row += 2
    
    # Long Term Liabilities (using long_term_liab which includes accounts moved from current)
    ws.cell(row=current_row, column=2, value='Long Term Liabilities').font = header_font
    current_row += 2
    
    # Group long-term liabilities (long_term_liab already includes moved accounts)
    floor_plan_accounts = []
    lease_accounts = []
    notes_accounts = []
    exec_lease_accounts = []
    other_lt_accounts = []
    
    for acc in long_term_liab:
        desc = acc['description'].upper()
        if 'FLOOR PLAN' in desc:
            floor_plan_accounts.append(acc)
        elif 'OPERATING LEASE' in desc or 'RIGHT OF USE' in desc:
            lease_accounts.append(acc)
        elif 'EXEC' in desc and 'LEASE' in desc:
            exec_lease_accounts.append(acc)
        elif 'NOTES PAYABLE' in desc or 'CONTRACTS PAYABLE' in desc:
            notes_accounts.append(acc)
        else:
            other_lt_accounts.append(acc)
    
    lt_total = 0
    
    # Main long-term debt (negate balance)
    main_lt_accounts = floor_plan_accounts + lease_accounts + notes_accounts + other_lt_accounts
    main_lt_total = 0
    for acc in sorted(main_lt_accounts, key=lambda x: x['account']):
        ws.cell(row=current_row, column=2, value=acc['description'])
        ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
        main_lt_total += -acc['balance']
        current_row += 1
    
    if main_lt_accounts:
        ws.cell(row=current_row, column=2, value=' Total')
        ws.cell(row=current_row, column=3, value=main_lt_total).number_format = money_format
        lt_total += main_lt_total
        current_row += 2
    
    # Executive Leases (negate balance)
    if exec_lease_accounts:
        ws.cell(row=current_row, column=2, value='Executive Leases').font = header_font
        current_row += 1
        exec_total = 0
        for acc in sorted(exec_lease_accounts, key=lambda x: x['account']):
            ws.cell(row=current_row, column=2, value=acc['description'])
            ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
            exec_total += -acc['balance']
            current_row += 1
        ws.cell(row=current_row, column=2, value='Executive Leases Total').font = header_font
        ws.cell(row=current_row, column=3, value=exec_total).number_format = money_format
        lt_total += exec_total
        current_row += 2
    
    ws.cell(row=current_row, column=2, value='Total Long Term Liabilities').font = header_font
    ws.cell(row=current_row, column=3, value=lt_total).number_format = money_format
    ws.cell(row=current_row, column=3).font = header_font
    current_row += 2
    
    # Stockholders Equity (negate balance - equity stored as negative credits)
    ws.cell(row=current_row, column=2, value='Stockholders Equity').font = header_font
    current_row += 2
    
    equity = bs_data['equity']
    equity_total = 0
    
    # Capital Stock
    for acc in sorted(equity['capital_stock'], key=lambda x: x['account']):
        ws.cell(row=current_row, column=2, value=acc['description'])
        ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
        equity_total += -acc['balance']
        current_row += 1
    
    # Distributions (keep as is - distributions reduce equity so sign is correct)
    for acc in sorted(equity['distributions'], key=lambda x: x['account']):
        ws.cell(row=current_row, column=2, value=acc['description'])
        ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
        equity_total += -acc['balance']
        current_row += 1
    
    # Retained Earnings
    for acc in sorted(equity['retained_earnings'], key=lambda x: x['account']):
        ws.cell(row=current_row, column=2, value=acc['description'])
        ws.cell(row=current_row, column=3, value=-acc['balance']).number_format = money_format
        equity_total += -acc['balance']
        current_row += 1
    
    current_row += 1
    ws.cell(row=current_row, column=2, value='Total Stockholders Equity').font = header_font
    ws.cell(row=current_row, column=3, value=equity_total).number_format = money_format
    ws.cell(row=current_row, column=3).font = header_font
    current_row += 2
    
    # Net Income YTD (calculated to balance the sheet)
    # Total Assets = Total Liabilities + Total Equity + Net Income YTD
    total_liabilities = current_liab_total + lt_total
    net_income_ytd = total_assets - total_liabilities - equity_total
    
    ws.cell(row=current_row, column=2, value='Net Income YTD')
    ws.cell(row=current_row, column=3, value=net_income_ytd).number_format = money_format
    current_row += 1
    
    # Total Liabilities and Stockholders' Equity
    total_liab_equity = total_liabilities + equity_total + net_income_ytd
    ws.cell(row=current_row, column=2, value="Total Liabilities and Stockholders' Equity").font = section_font
    ws.cell(row=current_row, column=3, value=total_liab_equity).number_format = money_format
    ws.cell(row=current_row, column=3).font = section_font
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    
    return ws


@pl_detailed_bp.route('/api/reports/pl/detailed/export', methods=['GET'])
@jwt_required()
def export_detailed_pl():
    """
    Export detailed P&L report to Excel with department tabs
    Matches the accounting firm's format exactly
    
    Query Parameters:
        year: Year for the report (default: current year)
        month: Month for the report (default: current month)
    """
    try:
        schema = get_tenant_schema()
        
        # Get parameters
        now = datetime.now()
        year = request.args.get('year', type=int, default=now.year)
        month = request.args.get('month', type=int, default=now.month)
        
        logger.info(f"Generating detailed P&L export for {year}-{month:02d}, schema: {schema}")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Get data for all departments
        all_dept_data = {}
        dept_order = ['new_equipment', 'used_equipment', 'parts', 'service', 'rental', 'transportation', 'in_house']
        
        total_gross_profit_mtd = 0
        total_gross_profit_ytd = 0
        
        for dept_key in dept_order:
            dept_data = get_department_detail(schema, dept_key, year, month)
            if dept_data:
                all_dept_data[dept_key] = dept_data
                # Don't create a separate worksheet for in_house - it's handled in the In House expenses tab
                if dept_key != 'in_house':
                    create_department_worksheet(wb, dept_data, year, month)
                total_gross_profit_mtd += dept_data['gross_profit_mtd']
                total_gross_profit_ytd += dept_data['gross_profit_ytd']
        
        # Get expense and other income data
        expense_data = get_overhead_expenses(schema, year, month)
        other_data = get_other_income_expense(schema, year, month)
        
        # Create In House worksheet with expenses
        inhouse_summary = {
            'total_gross_profit_mtd': total_gross_profit_mtd,
            'total_gross_profit_ytd': total_gross_profit_ytd
        }
        # Get the in_house department data for Sales/COGS section
        inhouse_dept_data = all_dept_data.get('in_house', None)
        create_inhouse_worksheet(wb, inhouse_summary, expense_data, other_data, year, month, inhouse_dept_data)
        
        # Create consolidated summary worksheet
        create_consolidated_worksheet(wb, all_dept_data, expense_data, other_data, year, month)
        
        # Create Balance Sheet worksheet
        create_balance_sheet_worksheet(wb, year, month)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        month_name = calendar.month_name[month]
        filename = f"ProfitLoss_Detailed_{month_name}{year}.xlsx"
        
        logger.info(f"Detailed P&L Excel export generated: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting detailed P&L Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to export detailed P&L to Excel', 'message': str(e)}), 500


@pl_detailed_bp.route('/api/reports/pl/detailed', methods=['GET'])
@jwt_required()
def get_detailed_pl():
    """
    Get detailed P&L data as JSON
    
    Query Parameters:
        year: Year for the report (default: current year)
        month: Month for the report (default: current month)
        department: Optional department key to filter
    """
    try:
        schema = get_tenant_schema()
        
        # Get parameters
        now = datetime.now()
        year = request.args.get('year', type=int, default=now.year)
        month = request.args.get('month', type=int, default=now.month)
        department = request.args.get('department', type=str, default=None)
        
        if department:
            # Return single department data
            dept_data = get_department_detail(schema, department, year, month)
            if dept_data:
                return jsonify(dept_data)
            else:
                return jsonify({'error': f'Department {department} not found'}), 404
        
        # Return all departments
        all_dept_data = {}
        dept_order = ['new_equipment', 'used_equipment', 'parts', 'service', 'rental', 'transportation', 'administrative']
        
        for dept_key in dept_order:
            dept_data = get_department_detail(schema, dept_key, year, month)
            if dept_data:
                all_dept_data[dept_key] = dept_data
        
        # Get expense and other income data
        expense_data = get_overhead_expenses(schema, year, month)
        other_data = get_other_income_expense(schema, year, month)
        
        return jsonify({
            'year': year,
            'month': month,
            'departments': all_dept_data,
            'overhead_expenses': expense_data,
            'other_income_expense': other_data
        })
        
    except Exception as e:
        logger.error(f"Error fetching detailed P&L: {str(e)}")
        return jsonify({'error': str(e)}), 500
