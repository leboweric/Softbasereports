"""
Currie Financial Model Report API
Automates quarterly Currie reporting by extracting data from Softbase
"""

from flask import Blueprint, jsonify, request
from flask_jwt_extended import jwt_required, get_jwt_identity
from src.utils.tenant_utils import get_tenant_db, get_tenant_schema
from src.services.cache_service import cache_service
from datetime import datetime, timedelta
import logging
import calendar

from flask_jwt_extended import get_jwt_identity
from src.models.user import User
from src.config.gl_accounts_loader import get_gl_accounts, get_currie_mappings

logger = logging.getLogger(__name__)

currie_bp = Blueprint('currie', __name__)
# sql_service is now obtained via get_tenant_db() for multi-tenant support
_sql_service = None
def get_sql_service():
    return get_tenant_db()
@currie_bp.route('/api/currie/sales-cogs-gp', methods=['GET'])
@jwt_required()
def get_sales_cogs_gp():
    """
    Get Sales, COGS, and Gross Profit data for Currie Financial Model
    Query params: start_date, end_date
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        force_refresh = request.args.get('refresh', 'false').lower() == 'true'
        
        if not start_date or not end_date:
            return jsonify({'error': 'start_date and end_date are required'}), 400
        
        # Use cache with 1-hour TTL, include tenant schema to isolate orgs
        user_identity = get_jwt_identity()
        schema = get_tenant_schema()
        cache_key = f'currie_sales_cogs_gp:{schema}:{start_date}:{end_date}'
        
        def fetch_data():
            return _fetch_sales_cogs_gp_data(start_date, end_date, user_identity)
        
        result = cache_service.cache_query(cache_key, fetch_data, ttl_seconds=3600, force_refresh=force_refresh)
        return jsonify(result), 200
        
    except Exception as e:
        logger.error(f"Error fetching Currie sales data: {str(e)}")
        return jsonify({'error': str(e)}), 500

def _fetch_sales_cogs_gp_data(start_date, end_date, user_identity):
    """Internal function to fetch Currie sales/COGS/GP data"""
    try:
        
        # Calculate number of months
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        months_diff = (end.year - start.year) * 12 + end.month - start.month + 1
        
        # Get organization name and tenant schema for the current user
        user = User.query.get(user_identity)
        org_name = user.organization.name if user and user.organization else 'Unknown'
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        
        # Get all revenue and COGS data
        data = {
            'dealership_info': {
                'name': org_name,
                'submitted_by': user_identity,
                'date': datetime.now().strftime('%Y-%m-%d'),
                'num_locations': 1,  # TODO: Make configurable
                'num_months': months_diff,
                'start_date': start_date,
                'end_date': end_date
            },
            'new_equipment': get_new_equipment_sales(start_date, end_date),
            'rental_data': get_rental_revenue(start_date, end_date),
            'service': get_service_revenue(start_date, end_date),
            'parts': get_parts_revenue(start_date, end_date),
            'trucking': get_trucking_revenue(start_date, end_date)
        }
        # Flatten rental_data for backward compatibility: data['rental'] = combined rental+rtr
        rental_result = data.pop('rental_data')
        data['rental'] = rental_result.get('rental', {'sales': 0, 'cogs': 0, 'gross_profit': 0})
        data['rtr'] = rental_result.get('rtr', {'sales': 0, 'cogs': 0, 'gross_profit': 0})
        
        # Calculate totals
        data['totals'] = calculate_totals(data, months_diff)
        
        # Get expenses for bottom summary section
        expenses = get_gl_expenses(start_date, end_date)
        data['expenses'] = expenses
        
        # Get other income and interest
        other_income_interest = get_other_income_and_interest(start_date, end_date)
        data['other_income'] = other_income_interest.get('other_income', 0)
        data['interest_expense'] = other_income_interest.get('interest_expense', 0)
        data['fi_income'] = other_income_interest.get('fi_income', 0)
        
        # Calculate bottom summary totals
        total_operating_profit = data['totals']['total_company']['gross_profit'] - expenses['grand_total'] + data['other_income'] + data['interest_expense']
        data['total_operating_profit'] = total_operating_profit
        data['pre_tax_income'] = total_operating_profit + data['fi_income']
        
        # Add department-allocated expenses for Expenses & Metrics tab
        dept_allocations = currie.get('dept_allocations', {
            'new': 0.47517,
            'used': 0.03209,
            'rental': 0.20694,
            'parts': 0.13121,
            'service': 0.14953,
            'trucking': 0.00507
        })
        
        personnel_total = expenses.get('personnel', {}).get('total', 0)
        operating_total = expenses.get('operating', {}).get('total', 0)
        occupancy_total = expenses.get('occupancy', {}).get('total', 0)
        
        # Calculate department allocations
        personnel_new = personnel_total * dept_allocations['new']
        personnel_used = personnel_total * dept_allocations['used']
        personnel_parts = personnel_total * dept_allocations['parts']
        personnel_service = personnel_total * dept_allocations['service']
        personnel_rental = personnel_total * dept_allocations['rental']
        personnel_trucking = personnel_total * dept_allocations['trucking']
        personnel_ga = personnel_total * (1 - sum(dept_allocations.values()))
        
        operating_new = operating_total * dept_allocations['new']
        operating_used = operating_total * dept_allocations['used']
        operating_parts = operating_total * dept_allocations['parts']
        operating_service = operating_total * dept_allocations['service']
        operating_rental = operating_total * dept_allocations['rental']
        operating_trucking = operating_total * dept_allocations['trucking']
        operating_ga = operating_total * (1 - sum(dept_allocations.values()))
        
        occupancy_new = occupancy_total * dept_allocations['new']
        occupancy_used = occupancy_total * dept_allocations['used']
        occupancy_parts = occupancy_total * dept_allocations['parts']
        occupancy_service = occupancy_total * dept_allocations['service']
        occupancy_rental = occupancy_total * dept_allocations['rental']
        occupancy_trucking = occupancy_total * dept_allocations['trucking']
        occupancy_ga = occupancy_total * (1 - sum(dept_allocations.values()))
        
        data['department_expenses'] = {
            'personnel': {
                'new': personnel_new,
                'used': personnel_used,
                'total_sales_dept': personnel_new + personnel_used,
                'parts': personnel_parts,
                'service': personnel_service,
                'rental': personnel_rental,
                'trucking': personnel_trucking,
                'ga': personnel_ga,
                'total': personnel_total
            },
            'operating': {
                'new': operating_new,
                'used': operating_used,
                'total_sales_dept': operating_new + operating_used,
                'parts': operating_parts,
                'service': operating_service,
                'rental': operating_rental,
                'trucking': operating_trucking,
                'ga': operating_ga,
                'total': operating_total
            },
            'occupancy': {
                'new': occupancy_new,
                'used': occupancy_used,
                'total_sales_dept': occupancy_new + occupancy_used,
                'parts': occupancy_parts,
                'service': occupancy_service,
                'rental': occupancy_rental,
                'trucking': occupancy_trucking,
                'ga': occupancy_ga,
                'total': occupancy_total
            },
            'total': {
                'new': personnel_new + operating_new + occupancy_new,
                'used': personnel_used + operating_used + occupancy_used,
                'total_sales_dept': (personnel_new + personnel_used) + (operating_new + operating_used) + (occupancy_new + occupancy_used),
                'parts': personnel_parts + operating_parts + occupancy_parts,
                'service': personnel_service + operating_service + occupancy_service,
                'rental': personnel_rental + operating_rental + occupancy_rental,
                'trucking': personnel_trucking + operating_trucking + occupancy_trucking,
                'ga': personnel_ga + operating_ga + occupancy_ga,
                'total': personnel_total + operating_total + occupancy_total
            }
        }
        
        # Add AR Aging data
        data['ar_aging'] = get_ar_aging()
        
        # Add Balance Sheet data
        data['balance_sheet'] = get_balance_sheet_data(end_date)
        
        return data
        
    except Exception as e:
        logger.error(f"Error fetching Currie sales data: {str(e)}")
        raise e


def get_new_equipment_sales(start_date, end_date):
    """Get new equipment sales broken down by category using GLDetail table"""
    try:
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        ne = currie.get('new_equipment', {})
        ue = currie.get('used_equipment', {})

        # Build account-to-category mapping from currie config
        revenue_map = {}  # account -> (category, 'sales')
        cost_map = {}     # account -> (category, 'cogs')
        
        for cat_key in ['new_lift_truck_primary', 'new_lift_truck_other', 'new_allied',
                        'batteries', 'other_new_equipment', 'operator_training', 'ecommerce', 'systems']:
            cat_config = ne.get(cat_key, {})
            for acct in cat_config.get('revenue', []):
                revenue_map[acct] = cat_key
            for acct in cat_config.get('cogs', []):
                cost_map[acct] = cat_key
        
        # Used equipment accounts
        for acct in ue.get('revenue', []):
            revenue_map[acct] = 'used_equipment'
        for acct in ue.get('cogs', []):
            cost_map[acct] = 'used_equipment'

        # Collect all accounts for the query
        all_accounts = list(set(list(revenue_map.keys()) + list(cost_map.keys())))
        if not all_accounts:
            return {}
        accounts_list = "', '".join(all_accounts)

        query = f"""
        SELECT 
            AccountNo,
            SUM(Amount) as total_amount
        FROM {schema}.GLDetail
        WHERE EffectiveDate >= %s 
          AND EffectiveDate <= %s
          AND Posted = 1
          AND AccountNo IN ('{accounts_list}')
        GROUP BY AccountNo
        """
        
        results = get_sql_service().execute_query(query, [start_date, end_date])
        
        # Initialize categories
        categories = {
            'new_lift_truck_primary': {'sales': 0, 'cogs': 0},
            'new_lift_truck_other': {'sales': 0, 'cogs': 0},
            'new_allied': {'sales': 0, 'cogs': 0},
            'other_new_equipment': {'sales': 0, 'cogs': 0},
            'operator_training': {'sales': 0, 'cogs': 0},
            'used_equipment': {'sales': 0, 'cogs': 0},
            'ecommerce': {'sales': 0, 'cogs': 0},
            'systems': {'sales': 0, 'cogs': 0},
            'batteries': {'sales': 0, 'cogs': 0}
        }
        
        # Map GL accounts to categories dynamically
        for row in results:
            account = str(row['AccountNo'])
            amount = float(row['total_amount'] or 0)
            
            if account in revenue_map:
                cat_key = revenue_map[account]
                categories[cat_key]['sales'] += -amount  # Revenue is credit (negative)
            elif account in cost_map:
                cat_key = cost_map[account]
                categories[cat_key]['cogs'] += amount  # COGS is debit (positive)
        
        # Calculate gross profit for each category
        for category in categories.values():
            category['gross_profit'] = category['sales'] - category['cogs']
        
        return categories
        
    except Exception as e:
        logger.error(f"Error fetching new equipment sales: {str(e)}")
        return {}


def get_rental_revenue(start_date, end_date):
    """Get rental revenue and RTR (Rental Truck Repair) as separate categories using GLDetail.
    Returns a dict with 'rental' and 'rtr' keys, each containing sales/cogs/gross_profit.
    For tenants without RTR config, rtr will be zeros."""
    try:
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        rental_cfg = currie.get('rental', {})
        rtr_cfg = currie.get('rtr', {})
        
        rental_rev = rental_cfg.get('revenue', [])
        rental_cos = rental_cfg.get('cogs', [])
        rtr_rev = rtr_cfg.get('revenue', [])
        rtr_cos = rtr_cfg.get('cogs', [])
        
        all_accounts = rental_rev + rental_cos + rtr_rev + rtr_cos
        if not all_accounts:
            return {
                'rental': {'sales': 0, 'cogs': 0, 'gross_profit': 0},
                'rtr': {'sales': 0, 'cogs': 0, 'gross_profit': 0}
            }
        accounts_list = "', '".join(all_accounts)

        query = f"""
        SELECT 
            AccountNo,
            SUM(Amount) as total_amount
        FROM {schema}.GLDetail
        WHERE EffectiveDate >= %s 
          AND EffectiveDate <= %s
          AND Posted = 1
          AND AccountNo IN ('{accounts_list}')
        GROUP BY AccountNo
        """
        
        results = get_sql_service().execute_query(query, [start_date, end_date])
        
        rental_data = {'sales': 0, 'cogs': 0, 'gross_profit': 0}
        rtr_data = {'sales': 0, 'cogs': 0, 'gross_profit': 0}
        
        for row in results:
            account = str(row['AccountNo'])
            amount = float(row['total_amount'] or 0)
            
            if account in rental_rev:
                rental_data['sales'] += -amount
            elif account in rental_cos:
                rental_data['cogs'] += amount
            elif account in rtr_rev:
                rtr_data['sales'] += -amount
            elif account in rtr_cos:
                rtr_data['cogs'] += amount
        
        rental_data['gross_profit'] = rental_data['sales'] - rental_data['cogs']
        rtr_data['gross_profit'] = rtr_data['sales'] - rtr_data['cogs']
        
        return {
            'rental': rental_data,
            'rtr': rtr_data
        }
        
    except Exception as e:
        logger.error(f"Error fetching rental revenue: {str(e)}")
        return {
            'rental': {'sales': 0, 'cogs': 0, 'gross_profit': 0},
            'rtr': {'sales': 0, 'cogs': 0, 'gross_profit': 0}
        }


def get_service_revenue(start_date, end_date):
    """Get service revenue broken down by customer, internal, warranty, sublet using GLDetail"""
    try:
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        svc = currie.get('service', {})

        # Build category mappings from config
        categories_config = {
            'customer_labor': svc.get('customer_labor', {}),
            'internal_labor': svc.get('internal_labor', {}),
            'warranty_labor': svc.get('warranty_labor', {}),
            'sublet': svc.get('sublet', {}),
            'other': svc.get('other', {})
        }

        # Build revenue and cost maps
        revenue_map = {}  # account -> category_key
        cost_map = {}     # account -> category_key
        all_accounts = []
        for cat_key, cat_cfg in categories_config.items():
            for acct in cat_cfg.get('revenue', []):
                revenue_map[acct] = cat_key
                all_accounts.append(acct)
            for acct in cat_cfg.get('cogs', []):
                cost_map[acct] = cat_key
                all_accounts.append(acct)

        if not all_accounts:
            return {}
        accounts_list = "', '".join(list(set(all_accounts)))

        query = f"""
        SELECT 
            AccountNo,
            SUM(Amount) as total_amount
        FROM {schema}.GLDetail
        WHERE EffectiveDate >= %s 
          AND EffectiveDate <= %s
          AND Posted = 1
          AND AccountNo IN ('{accounts_list}')
        GROUP BY AccountNo
        """
        
        results = get_sql_service().execute_query(query, [start_date, end_date])
        
        service_data = {
            'customer_labor': {'sales': 0, 'cogs': 0},
            'internal_labor': {'sales': 0, 'cogs': 0},
            'warranty_labor': {'sales': 0, 'cogs': 0},
            'sublet': {'sales': 0, 'cogs': 0},
            'other': {'sales': 0, 'cogs': 0}
        }
        
        for row in results:
            account = str(row['AccountNo'])
            amount = float(row['total_amount'] or 0)
            
            if account in revenue_map:
                cat_key = revenue_map[account]
                service_data[cat_key]['sales'] += -amount  # Revenue is credit (negative)
            elif account in cost_map:
                cat_key = cost_map[account]
                service_data[cat_key]['cogs'] += amount  # COGS is debit (positive)
        
        # Calculate gross profit
        for category in service_data.values():
            category['gross_profit'] = category['sales'] - category['cogs']
        
        return service_data
        
    except Exception as e:
        logger.error(f"Error fetching service revenue: {str(e)}")
        return {}


def get_parts_revenue(start_date, end_date):
    """Get parts revenue broken down by counter, RO, internal, warranty using GLDetail"""
    try:
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        parts_cfg = currie.get('parts', {})

        # Build category mappings from config
        categories_config = {
            'counter_primary': parts_cfg.get('counter_primary', {}),
            'counter_other': parts_cfg.get('counter_other', {}),
            'ro_primary': parts_cfg.get('ro_primary', {}),
            'ro_other': parts_cfg.get('ro_other', {}),
            'internal': parts_cfg.get('internal', {}),
            'warranty': parts_cfg.get('warranty', {}),
            'ecommerce': parts_cfg.get('ecommerce', {})
        }

        # Build revenue and cost maps
        revenue_map = {}
        cost_map = {}
        all_accounts = []
        for cat_key, cat_cfg in categories_config.items():
            for acct in cat_cfg.get('revenue', []):
                revenue_map[acct] = cat_key
                all_accounts.append(acct)
            for acct in cat_cfg.get('cogs', []):
                cost_map[acct] = cat_key
                all_accounts.append(acct)

        if not all_accounts:
            return {}
        accounts_list = "', '".join(list(set(all_accounts)))

        query = f"""
        SELECT 
            AccountNo,
            SUM(Amount) as total_amount
        FROM {schema}.GLDetail
        WHERE EffectiveDate >= %s 
          AND EffectiveDate <= %s
          AND Posted = 1
          AND AccountNo IN ('{accounts_list}')
        GROUP BY AccountNo
        """
        
        results = get_sql_service().execute_query(query, [start_date, end_date])
        
        parts_data = {
            'counter_primary': {'sales': 0, 'cogs': 0},
            'counter_other': {'sales': 0, 'cogs': 0},
            'ro_primary': {'sales': 0, 'cogs': 0},
            'ro_other': {'sales': 0, 'cogs': 0},
            'internal': {'sales': 0, 'cogs': 0},
            'warranty': {'sales': 0, 'cogs': 0},
            'ecommerce': {'sales': 0, 'cogs': 0}
        }
        
        for row in results:
            account = str(row['AccountNo'])
            amount = float(row['total_amount'] or 0)
            
            if account in revenue_map:
                cat_key = revenue_map[account]
                parts_data[cat_key]['sales'] += -amount  # Revenue is credit (negative)
            elif account in cost_map:
                cat_key = cost_map[account]
                parts_data[cat_key]['cogs'] += amount  # COGS is debit (positive)
        
        # Calculate gross profit
        for category in parts_data.values():
            category['gross_profit'] = category['sales'] - category['cogs']
        
        return parts_data
        
    except Exception as e:
        logger.error(f"Error fetching parts revenue: {str(e)}")
        return {}


def get_trucking_revenue(start_date, end_date):
    """Get trucking/delivery revenue using GLDetail with all trucking GL accounts"""
    try:
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        trucking_cfg = currie.get('trucking', {})
        
        revenue_accounts = trucking_cfg.get('revenue', [])
        cost_accounts = trucking_cfg.get('cogs', [])
        all_accounts = revenue_accounts + cost_accounts
        if not all_accounts:
            return {'sales': 0, 'cogs': 0, 'gross_profit': 0}
        accounts_list = "', '".join(all_accounts)

        query = f"""
        SELECT 
            AccountNo,
            SUM(Amount) as total_amount
        FROM {schema}.GLDetail
        WHERE EffectiveDate >= %s 
          AND EffectiveDate <= %s
          AND Posted = 1
          AND AccountNo IN ('{accounts_list}')
        GROUP BY AccountNo
        """
        
        results = get_sql_service().execute_query(query, [start_date, end_date])
        
        trucking_data = {'sales': 0, 'cogs': 0, 'gross_profit': 0}
        
        for row in results:
            account = str(row['AccountNo'])
            amount = float(row['total_amount'] or 0)
            
            if account in revenue_accounts:
                trucking_data['sales'] += -amount  # Revenue is credit (negative)
            elif account in cost_accounts:
                trucking_data['cogs'] += amount  # COGS is debit (positive)
        
        trucking_data['gross_profit'] = trucking_data['sales'] - trucking_data['cogs']
        
        return trucking_data
        
    except Exception as e:
        logger.error(f"Error fetching trucking revenue: {str(e)}")
        return {}


def calculate_totals(data, num_months):
    """Calculate total sales, COGS, and GP across all categories with subtotals"""
    
    # Calculate subtotals for each section
    # TOTAL NEW EQUIPMENT = only first 4 items (matches Excel row 14)
    total_new_equipment = {'sales': 0, 'cogs': 0, 'gross_profit': 0}
    if 'new_equipment' in data:
        # Only sum the first 4 equipment categories for TOTAL NEW EQUIPMENT subtotal
        new_eq_categories = ['new_lift_truck_primary', 'new_lift_truck_other', 'new_allied', 'other_new_equipment']
        for cat in new_eq_categories:
            if cat in data['new_equipment']:
                total_new_equipment['sales'] += data['new_equipment'][cat].get('sales', 0)
                total_new_equipment['cogs'] += data['new_equipment'][cat].get('cogs', 0)
        total_new_equipment['gross_profit'] = total_new_equipment['sales'] - total_new_equipment['cogs']
    
    # TOTAL SALES DEPT = all equipment items (matches Excel row 20)
    total_sales_dept = {'sales': 0, 'cogs': 0, 'gross_profit': 0}
    if 'new_equipment' in data:
        for category in data['new_equipment'].values():
            total_sales_dept['sales'] += category.get('sales', 0)
            total_sales_dept['cogs'] += category.get('cogs', 0)
        total_sales_dept['gross_profit'] = total_sales_dept['sales'] - total_sales_dept['cogs']
    
    total_rental = {'sales': 0, 'cogs': 0, 'gross_profit': 0}
    if 'rental' in data:
        total_rental['sales'] = data['rental'].get('sales', 0)
        total_rental['cogs'] = data['rental'].get('cogs', 0)
    # Include RTR in total rental department
    if 'rtr' in data:
        total_rental['sales'] += data['rtr'].get('sales', 0)
        total_rental['cogs'] += data['rtr'].get('cogs', 0)
    total_rental['gross_profit'] = total_rental['sales'] - total_rental['cogs']
    
    total_service = {'sales': 0, 'cogs': 0, 'gross_profit': 0}
    if 'service' in data:
        for category in data['service'].values():
            total_service['sales'] += category.get('sales', 0)
            total_service['cogs'] += category.get('cogs', 0)
        total_service['gross_profit'] = total_service['sales'] - total_service['cogs']
    
    total_parts = {'sales': 0, 'cogs': 0, 'gross_profit': 0}
    if 'parts' in data:
        for category in data['parts'].values():
            total_parts['sales'] += category.get('sales', 0)
            total_parts['cogs'] += category.get('cogs', 0)
        total_parts['gross_profit'] = total_parts['sales'] - total_parts['cogs']
    
    # Calculate combined totals
    total_aftermarket = {
        'sales': total_service['sales'] + total_parts['sales'],
        'cogs': total_service['cogs'] + total_parts['cogs'],
        'gross_profit': 0
    }
    total_aftermarket['gross_profit'] = total_aftermarket['sales'] - total_aftermarket['cogs']
    
    # Grand total uses total_sales_dept (all equipment) not just total_new_equipment (first 4)
    grand_total = {
        'sales': total_sales_dept['sales'] + total_rental['sales'] + total_service['sales'] + total_parts['sales'],
        'cogs': total_sales_dept['cogs'] + total_rental['cogs'] + total_service['cogs'] + total_parts['cogs'],
        'gross_profit': 0
    }
    
    # Add trucking to grand total
    if 'trucking' in data:
        grand_total['sales'] += data['trucking'].get('sales', 0)
        grand_total['cogs'] += data['trucking'].get('cogs', 0)
    
    grand_total['gross_profit'] = grand_total['sales'] - grand_total['cogs']
    
    # Calculate average monthly sales & GP
    avg_monthly_sales_gp = grand_total['sales'] / num_months if num_months > 0 else 0
    
    return {
        'total_new_equipment': total_new_equipment,  # Subtotal for first 4 items only (Excel row 14)
        'total_sales_dept': total_sales_dept,        # Grand total for all equipment (Excel row 20)
        'total_rental': total_rental,
        'total_service': total_service,
        'total_parts': total_parts,
        'total_aftermarket': total_aftermarket,
        'total_net_sales_gp': grand_total,  # Same as total_company but matches Excel naming
        'grand_total': grand_total,
        'total_company': grand_total,  # Alias for grand_total
        'avg_monthly_sales_gp': avg_monthly_sales_gp,
        # Keep legacy format for backward compatibility
        'sales': grand_total['sales'],
        'cogs': grand_total['cogs'],
        'gross_profit': grand_total['gross_profit']
    }


@currie_bp.route('/api/currie/metrics', methods=['GET'])
@jwt_required()
def get_currie_metrics():
    """Get metrics for Currie Financial Model"""
    try:
        # Get date range from query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        force_refresh = request.args.get('refresh', 'false').lower() == 'true'
        
        if not start_date or not end_date:
            return jsonify({'error': 'start_date and end_date are required'}), 400
        
        # Use cache with 1-hour TTL
        schema = get_tenant_schema()
        cache_key = f'currie_metrics:{schema}:{start_date}:{end_date}'
        
        def fetch_metrics():
            return _fetch_currie_metrics_data(start_date, end_date)
        
        result = cache_service.cache_query(cache_key, fetch_metrics, ttl_seconds=3600, force_refresh=force_refresh)
        return jsonify(result), 200
        
    except Exception as e:
        logger.error(f"Error fetching Currie metrics: {str(e)}")
        return jsonify({'error': 'Failed to fetch metrics', 'message': str(e)}), 500

def _fetch_currie_metrics_data(start_date, end_date):
    """Internal function to fetch Currie metrics data"""
    # Calculate number of days in period
    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')
    num_days = (end - start).days + 1
    
    metrics = {}
    
    # 1. AR Aging
    metrics['ar_aging'] = get_ar_aging()
    
    # 2. Service Calls Per Day
    metrics['service_calls_per_day'] = get_service_calls_per_day(start_date, end_date, num_days)
    
    # 3. Technician Count
    metrics['technician_count'] = get_technician_count(start_date, end_date)
    
    # 4. Labor Metrics
    metrics['labor_metrics'] = get_labor_metrics(start_date, end_date)
    
    # 5. Parts Inventory Metrics
    metrics['parts_inventory'] = get_parts_inventory_metrics(start_date, end_date)
    
    # 6. Absorption Rate
    # Get revenue and expense data to calculate absorption rate
    num_months = (end - start).days / 30.44  # Average days per month
    rental_result = get_rental_revenue(start_date, end_date)
    data = {
        'rental': rental_result.get('rental', {'sales': 0, 'cogs': 0, 'gross_profit': 0}),
        'rtr': rental_result.get('rtr', {'sales': 0, 'cogs': 0, 'gross_profit': 0}),
        'service': get_service_revenue(start_date, end_date),
        'parts': get_parts_revenue(start_date, end_date)
    }
    totals = calculate_totals(data, max(1, int(num_months)))
    expenses = get_gl_expenses(start_date, end_date)
    
    # Absorption Rate = (Aftermarket GP / Total Expenses) × 100
    aftermarket_gp = totals.get('total_aftermarket', {}).get('gross_profit', 0)
    rental_gp = totals.get('total_rental', {}).get('gross_profit', 0)
    total_aftermarket_gp = aftermarket_gp + rental_gp
    
    total_expenses = (
        expenses.get('personnel', {}).get('total', 0) +
        expenses.get('operating', {}).get('total', 0) +
        expenses.get('occupancy', {}).get('total', 0)
    )
    
    absorption_rate = (total_aftermarket_gp / total_expenses * 100) if total_expenses > 0 else 0
    
    metrics['absorption_rate'] = {
        'rate': round(absorption_rate, 1),
        'aftermarket_gp': round(total_aftermarket_gp, 2),
        'total_expenses': round(total_expenses, 2)
    }
    
    # 7. Rental Fleet Metrics
    try:
        metrics['rental_fleet'] = get_rental_fleet_metrics(start_date, end_date)
    except Exception as e:
        logger.error(f"Error fetching rental fleet metrics: {str(e)}")
        metrics['rental_fleet'] = {}
    
    # 8. Revenue/GP per Technician (combine service revenue with tech count)
    tech_count = metrics.get('technician_count', {}).get('active_technicians', 0)
    service_revenue = totals.get('total_service', {}).get('sales', 0)
    service_gp_val = totals.get('total_service', {}).get('gross_profit', 0)
    num_months_calc = max(1, int(num_months))
    metrics['service_productivity'] = {
        'revenue_per_tech_monthly': round(service_revenue / num_months_calc / tech_count, 2) if tech_count > 0 else 0,
        'gp_per_tech_monthly': round(service_gp_val / num_months_calc / tech_count, 2) if tech_count > 0 else 0,
        'total_service_revenue': round(service_revenue, 2),
        'total_service_gp': round(service_gp_val, 2),
        'hours_per_tech_monthly': round((metrics.get('labor_metrics', {}).get('total_billed_hours', 0)) / num_months_calc / tech_count, 1) if tech_count > 0 else 0,
    }
    
    # 9. Department GP% summary for benchmarking
    service_total = totals.get('total_service', totals.get('total_aftermarket', {}))
    parts_total = totals.get('total_parts', {})
    rental_total = totals.get('total_rental', {})
    
    def calc_gp_pct(dept_data):
        sales = dept_data.get('sales', 0)
        gp = dept_data.get('gross_profit', 0)
        return round((gp / sales * 100) if sales > 0 else 0, 1)
    
    # Get service totals from the raw data
    svc_sales = sum(data.get('service', {}).get(k, {}).get('sales', 0) for k in ['customer_labor', 'internal_labor', 'warranty_labor', 'sublet', 'other'])
    svc_gp = sum(data.get('service', {}).get(k, {}).get('gross_profit', 0) for k in ['customer_labor', 'internal_labor', 'warranty_labor', 'sublet', 'other'])
    parts_sales = sum(data.get('parts', {}).get(k, {}).get('sales', 0) for k in data.get('parts', {}).keys())
    parts_gp_val = sum(data.get('parts', {}).get(k, {}).get('gross_profit', 0) for k in data.get('parts', {}).keys())
    # Combine rental + RTR for department benchmarks
    rental_sales = data.get('rental', {}).get('sales', 0) + data.get('rtr', {}).get('sales', 0)
    rental_gp_val = data.get('rental', {}).get('gross_profit', 0) + data.get('rtr', {}).get('gross_profit', 0)
    
    metrics['dept_gp_benchmarks'] = {
        'service': {
            'gp_pct': round((svc_gp / svc_sales * 100) if svc_sales > 0 else 0, 1),
            'target': 65.0,
            'sales': round(svc_sales, 2),
            'gp': round(svc_gp, 2)
        },
        'parts': {
            'gp_pct': round((parts_gp_val / parts_sales * 100) if parts_sales > 0 else 0, 1),
            'target': 40.0,
            'sales': round(parts_sales, 2),
            'gp': round(parts_gp_val, 2)
        },
        'rental': {
            'gp_pct': round((rental_gp_val / rental_sales * 100) if rental_sales > 0 else 0, 1),
            'target': 45.0,
            'sales': round(rental_sales, 2),
            'gp': round(rental_gp_val, 2)
        }
    }
    
    return {
        'metrics': metrics,
        'date_range': {
            'start_date': start_date,
            'end_date': end_date,
            'num_days': num_days
        },
        'generated_at': datetime.now().isoformat()
    }


def get_ar_aging():
    """Get AR aging buckets (reusing logic from department_reports)"""
    try:
        schema = get_tenant_schema()
        
        # Get total AR
        total_ar_query = f"""
        SELECT SUM(Amount) as total_ar
        FROM {schema}.ARDetail
        WHERE (HistoryFlag IS NULL OR HistoryFlag = 0)
            AND DeletionTime IS NULL
        """
        total_ar_result = get_sql_service().execute_query(total_ar_query, [])
        total_ar = float(total_ar_result[0]['total_ar']) if total_ar_result and total_ar_result[0]['total_ar'] else 0
        
        # Get AR aging buckets
        ar_query = f"""
        WITH InvoiceBalances AS (
            SELECT 
                ar.InvoiceNo,
                ar.CustomerNo,
                MIN(ar.Due) as Due,
                SUM(ar.Amount) as NetBalance
            FROM {schema}.ARDetail ar
            WHERE (ar.HistoryFlag IS NULL OR ar.HistoryFlag = 0)
                AND ar.DeletionTime IS NULL
                AND ar.InvoiceNo IS NOT NULL
            GROUP BY ar.InvoiceNo, ar.CustomerNo
            HAVING SUM(ar.Amount) > 0.01
        )
        SELECT 
            CASE 
                WHEN Due IS NULL THEN 'No Due Date'
                WHEN DATEDIFF(day, Due, GETDATE()) < 30 THEN 'Current'
                WHEN DATEDIFF(day, Due, GETDATE()) BETWEEN 30 AND 59 THEN '30-60'
                WHEN DATEDIFF(day, Due, GETDATE()) BETWEEN 60 AND 89 THEN '60-90'
                WHEN DATEDIFF(day, Due, GETDATE()) >= 90 THEN '90+'
            END as AgingBucket,
            SUM(NetBalance) as TotalAmount
        FROM InvoiceBalances
        GROUP BY 
            CASE 
                WHEN Due IS NULL THEN 'No Due Date'
                WHEN DATEDIFF(day, Due, GETDATE()) < 30 THEN 'Current'
                WHEN DATEDIFF(day, Due, GETDATE()) BETWEEN 30 AND 59 THEN '30-60'
                WHEN DATEDIFF(day, Due, GETDATE()) BETWEEN 60 AND 89 THEN '60-90'
                WHEN DATEDIFF(day, Due, GETDATE()) >= 90 THEN '90+'
            END
        """
        
        ar_results = get_sql_service().execute_query(ar_query, [])
        
        # Format results for Currie (Current, 31-60, 61-90, 91+)
        ar_aging = {
            'current': 0,
            'days_31_60': 0,
            'days_61_90': 0,
            'days_91_plus': 0,
            'total': total_ar
        }
        
        for row in ar_results:
            bucket = row['AgingBucket']
            amount = float(row['TotalAmount'] or 0)
            
            if bucket == 'Current':
                ar_aging['current'] = amount
            elif bucket == '30-60':
                ar_aging['days_31_60'] = amount
            elif bucket == '60-90':
                ar_aging['days_61_90'] = amount
            elif bucket == '90+':
                ar_aging['days_91_plus'] = amount
        
        return ar_aging
        
    except Exception as e:
        logger.error(f"Error fetching AR aging: {str(e)}")
        return {}


def get_service_calls_per_day(start_date, end_date, num_days):
    """Calculate average service calls per day"""
    try:
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        dept_codes = currie.get('service_dept_codes', [4, 5, 6, 7, 8])
        # Convert to SaleDept format (multiply by 10 for WO table format, or use as strings)
        # SaleDept codes are tenant-specific (loaded from currie mappings)
        dept_placeholders = ','.join(['%s'] * len(dept_codes))

        query = f"""
        SELECT COUNT(*) as total_service_calls
        FROM {schema}.WO
        WHERE OpenDate >= %s 
          AND OpenDate <= %s
          AND SaleDept IN ({dept_placeholders})
        """
        
        # Convert dept codes to strings for SaleDept column format
        dept_code_strs = [str(code) for code in dept_codes]
        results = get_sql_service().execute_query(query, [start_date, end_date] + dept_code_strs)
        
        total_calls = int(results[0]['total_service_calls']) if results and results[0]['total_service_calls'] else 0
        calls_per_day = total_calls / num_days if num_days > 0 else 0
        
        return {
            'total_service_calls': total_calls,
            'calls_per_day': round(calls_per_day, 2),
            'num_days': num_days
        }
        
    except Exception as e:
        logger.error(f"Error calculating service calls per day: {str(e)}")
        return {}


def get_technician_count(start_date, end_date):
    """Count unique technicians who worked during the period"""
    try:
        schema = get_tenant_schema()

        query = f"""
        SELECT COUNT(DISTINCT Technician) as technician_count
        FROM {schema}.WO
        WHERE OpenDate >= %s 
          AND OpenDate <= %s
          AND Technician IS NOT NULL
          AND Technician != ''
        """
        
        results = get_sql_service().execute_query(query, [start_date, end_date])
        
        count = int(results[0]['technician_count']) if results and results[0]['technician_count'] else 0
        
        return {
            'active_technicians': count
        }
        
    except Exception as e:
        logger.error(f"Error counting technicians: {str(e)}")
        return {}


def get_parts_inventory_metrics(start_date, end_date):
    """Get parts inventory metrics: fill rate, turnover, aging"""
    try:
        schema = get_tenant_schema()
        
        # Calculate days in period
        from datetime import datetime
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        days_in_period = (end - start).days + 1
        
        # 1. Fill Rate - percentage of orders filled from stock
        fill_rate_query = f"""
        WITH PartsOrders AS (
            SELECT 
                wp.PartNo,
                wp.BOQty as BackorderQty,
                COALESCE(p.OnHand, 0) as CurrentStock,
                CASE 
                    WHEN wp.BOQty > 0 THEN 'Backordered'
                    WHEN p.OnHand IS NULL OR p.OnHand = 0 THEN 'Out of Stock'
                    WHEN p.OnHand < wp.Qty THEN 'Partial Stock'
                    ELSE 'In Stock'
                END as StockStatus
            FROM {schema}.WOParts wp
            INNER JOIN {schema}.WO w ON wp.WONo = w.WONo
            LEFT JOIN {schema}.Parts p ON wp.PartNo = p.PartNo
            WHERE w.OpenDate >= %s AND w.OpenDate <= %s
        )
        SELECT 
            COUNT(*) as TotalOrders,
            SUM(CASE WHEN StockStatus = 'In Stock' THEN 1 ELSE 0 END) as FilledOrders
        FROM PartsOrders
        """
        
        fill_rate_result = get_sql_service().execute_query(fill_rate_query, [start_date, end_date])
        
        total_orders = int(fill_rate_result[0]['TotalOrders'] or 0) if fill_rate_result else 0
        filled_orders = int(fill_rate_result[0]['FilledOrders'] or 0) if fill_rate_result else 0
        fill_rate = (filled_orders / total_orders * 100) if total_orders > 0 else 0
        
        # 2. Inventory Value - TOTAL current inventory (not period-specific)
        inventory_query = f"""
        SELECT SUM(OnHand * Cost) as TotalInventoryValue
        FROM {schema}.Parts
        WHERE OnHand > 0 AND Cost > 0
        """
        
        inventory_result = get_sql_service().execute_query(inventory_query, [])
        inventory_value = float(inventory_result[0]['TotalInventoryValue'] or 0) if inventory_result else 0
        
        # 3. Inventory Turnover - annualized based on period movement
        turnover_query = f"""
        SELECT 
            SUM(wp.Qty * p.Cost) as TotalCOGS
        FROM {schema}.WOParts wp
        INNER JOIN {schema}.WO w ON wp.WONo = w.WONo
        LEFT JOIN {schema}.Parts p ON wp.PartNo = p.PartNo
        WHERE w.OpenDate >= %s AND w.OpenDate <= %s
        """
        
        turnover_result = get_sql_service().execute_query(turnover_query, [start_date, end_date])
        cogs = float(turnover_result[0]['TotalCOGS'] or 0) if turnover_result else 0
        
        # Annualize the turnover (COGS for period / avg inventory) * (365 / days in period)
        turnover_rate = (cogs / inventory_value * (365 / days_in_period)) if (inventory_value > 0 and days_in_period > 0) else 0
        
        # 4. Inventory Aging - parts with no movement in 90+ days
        aging_query = f"""
        WITH PartMovement AS (
            SELECT 
                p.PartNo,
                MAX(p.OnHand) as CurrentStock,
                MAX(p.Cost) as Cost,
                MAX(w.OpenDate) as LastMovementDate
            FROM {schema}.Parts p
            LEFT JOIN {schema}.WOParts wp ON p.PartNo = wp.PartNo
            LEFT JOIN {schema}.WO w ON wp.WONo = w.WONo
            WHERE p.OnHand > 0
            GROUP BY p.PartNo
        )
        SELECT 
            COUNT(CASE WHEN DATEDIFF(day, LastMovementDate, GETDATE()) > 365 OR LastMovementDate IS NULL THEN 1 END) as Obsolete,
            COUNT(CASE WHEN DATEDIFF(day, LastMovementDate, GETDATE()) BETWEEN 181 AND 365 THEN 1 END) as Slow,
            COUNT(CASE WHEN DATEDIFF(day, LastMovementDate, GETDATE()) BETWEEN 91 AND 180 THEN 1 END) as Medium,
            COUNT(CASE WHEN DATEDIFF(day, LastMovementDate, GETDATE()) <= 90 THEN 1 END) as Fast,
            SUM(CASE WHEN DATEDIFF(day, LastMovementDate, GETDATE()) > 365 OR LastMovementDate IS NULL THEN CurrentStock * Cost ELSE 0 END) as ObsoleteValue
        FROM PartMovement
        """
        
        aging_result = get_sql_service().execute_query(aging_query, [])
        
        if aging_result and len(aging_result) > 0:
            aging = aging_result[0]
            return {
                'fill_rate': round(fill_rate, 1),
                'total_orders': total_orders,
                'filled_orders': filled_orders,
                'inventory_turnover': round(turnover_rate, 2),
                'inventory_value': round(inventory_value, 2),
                'aging': {
                    'obsolete_count': int(aging.get('Obsolete', 0)),
                    'slow_count': int(aging.get('Slow', 0)),
                    'medium_count': int(aging.get('Medium', 0)),
                    'fast_count': int(aging.get('Fast', 0)),
                    'obsolete_value': round(float(aging.get('ObsoleteValue', 0)), 2)
                }
            }
        
        return {}
        
    except Exception as e:
        logger.error(f"Error fetching parts inventory metrics: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}


def get_rental_fleet_metrics(start_date, end_date):
    """Get rental fleet metrics: fleet value, depreciation, financial utilization"""
    try:
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        rental_fleet_bs = currie.get('rental_fleet_bs', {})
        gross_acct = rental_fleet_bs.get('gross_equipment', '183000')
        deprec_acct = rental_fleet_bs.get('accumulated_depreciation', '193000')
        
        # 1. Rental Fleet Value from GL accounts (use YTD from latest period)
        gross_like = gross_acct[:3] + '%'  # e.g. '183%'
        deprec_like = deprec_acct[:3] + '%'  # e.g. '193%'
        fleet_query = f"""
        SELECT 
            SUM(CASE WHEN AccountNo LIKE %s THEN YTD ELSE 0 END) as gross_fleet_value,
            SUM(CASE WHEN AccountNo LIKE %s THEN ABS(YTD) ELSE 0 END) as accumulated_depreciation
        FROM {schema}.GL
        WHERE (AccountNo LIKE %s OR AccountNo LIKE %s)
          AND Year = (SELECT MAX(Year) FROM {schema}.GL WHERE AccountNo LIKE %s)
          AND Month = (SELECT MAX(Month) FROM {schema}.GL WHERE AccountNo LIKE %s AND Year = (SELECT MAX(Year) FROM {schema}.GL WHERE AccountNo LIKE %s))
        """
        
        fleet_result = get_sql_service().execute_query(fleet_query, [gross_like, deprec_like, gross_like, deprec_like, gross_like, deprec_like, gross_like])
        gross_value = float(fleet_result[0]['gross_fleet_value'] or 0) if fleet_result else 0
        accum_deprec = float(fleet_result[0]['accumulated_depreciation'] or 0) if fleet_result else 0
        net_value = gross_value - accum_deprec
        
        # 2. Rental Revenue for the period (from GL)
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        days_in_period = (end - start).days + 1
        
        rental_accounts = currie.get('rental', {}).get('revenue', [])
        if rental_accounts:
            rental_placeholders = ','.join(['%s'] * len(rental_accounts))
        else:
            rental_placeholders = "'NONE'"
            rental_accounts = []
        rental_rev_query = f"""
        SELECT ABS(SUM(Amount)) as rental_revenue
        FROM {schema}.GLDetail
        WHERE AccountNo IN ({rental_placeholders})
          AND PostDate >= %s AND PostDate <= %s
          AND Posted = 1
        """
        
        rental_rev_result = get_sql_service().execute_query(rental_rev_query, rental_accounts + [start_date, end_date])
        period_revenue = float(rental_rev_result[0]['rental_revenue'] or 0) if rental_rev_result else 0
        
        # Annualize the revenue
        annualized_revenue = (period_revenue / days_in_period * 365) if days_in_period > 0 else 0
        
        # Financial Utilization = Annualized Revenue / Gross Fleet Value (acquisition cost)
        financial_utilization = (annualized_revenue / gross_value * 100) if gross_value > 0 else 0
        
        # Depreciation for the period
        # Depreciation account - derive LIKE pattern from the accumulated depreciation account prefix
        # Convention: accumulated depreciation prefix (e.g. 193) -> expense prefix (e.g. 593)
        deprec_expense_like = '5' + deprec_acct[1:3] + '%'
        deprec_query = f"""
        SELECT ABS(SUM(Amount)) as depreciation_expense
        FROM {schema}.GLDetail
        WHERE AccountNo LIKE %s
          AND PostDate >= %s AND PostDate <= %s
          AND Posted = 1
        """
        
        deprec_result = get_sql_service().execute_query(deprec_query, [deprec_expense_like, start_date, end_date])
        period_depreciation = float(deprec_result[0]['depreciation_expense'] or 0) if deprec_result else 0
        annualized_depreciation = (period_depreciation / days_in_period * 365) if days_in_period > 0 else 0
        
        # Rental Multiple = Revenue / Depreciation
        rental_multiple = (annualized_revenue / annualized_depreciation) if annualized_depreciation > 0 else 0
        
        # Fleet unit count from Equipment table
        unit_query = f"""
        SELECT COUNT(*) as unit_count
        FROM {schema}.Equipment
        WHERE Status = 'A'
          AND RentalUnit = 1
        """
        
        unit_result = get_sql_service().execute_query(unit_query, [])
        unit_count = int(unit_result[0]['unit_count'] or 0) if unit_result else 0
        
        return {
            'gross_fleet_value': round(gross_value, 2),
            'accumulated_depreciation': round(accum_deprec, 2),
            'net_fleet_value': round(net_value, 2),
            'unit_count': unit_count,
            'annualized_revenue': round(annualized_revenue, 2),
            'financial_utilization': round(financial_utilization, 1),
            'annualized_depreciation': round(annualized_depreciation, 2),
            'rental_multiple': round(rental_multiple, 2),
            'revenue_per_unit': round(annualized_revenue / unit_count, 2) if unit_count > 0 else 0
        }
        
    except Exception as e:
        logger.error(f"Error fetching rental fleet metrics: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}


def get_labor_metrics(start_date, end_date):
    """Get labor productivity metrics from WOLabor"""
    try:
        # Use the same pattern as other successful queries: SUM(Sell) for labor value
        schema = get_tenant_schema()

        query = f"""
        SELECT 
            COUNT(DISTINCT l.WONo) as wo_count,
            SUM(l.Hours) as total_hours,
            CASE 
                WHEN SUM(l.Hours) > 0 THEN SUM(l.Sell) / SUM(l.Hours)
                ELSE 0 
            END as avg_rate,
            SUM(l.Sell) as total_labor_value
        FROM {schema}.WOLabor l
        INNER JOIN {schema}.WO w ON l.WONo = w.WONo
        WHERE w.OpenDate >= %s 
          AND w.OpenDate <= %s
        """
        
        results = get_sql_service().execute_query(query, [start_date, end_date])
        
        if results and len(results) > 0:
            row = results[0]
            return {
                'work_orders_with_labor': int(row['wo_count'] or 0),
                'total_billed_hours': float(row['total_hours'] or 0),
                'average_labor_rate': float(row['avg_rate'] or 0),
                'total_labor_value': float(row['total_labor_value'] or 0)
            }
        
        return {}
        
    except Exception as e:
        logger.error(f"Error fetching labor metrics: {str(e)}")
        return {}


@currie_bp.route('/api/currie/export-excel', methods=['GET'])
@jwt_required()
def export_currie_excel():
    """Export Currie Financial Model to Excel using template"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from flask import send_file
        import os
        from datetime import datetime
        import io
        
        # Get date range from query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({'error': 'start_date and end_date are required'}), 400
        
        # Calculate number of months and days
        from datetime import datetime
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        months_diff = (end.year - start.year) * 12 + (end.month - start.month) + 1
        num_days = (end - start).days + 1
        
        # Get tenant schema for currie mappings
        schema = get_tenant_schema()
        
        # Get all the data
        new_equipment = get_new_equipment_sales(start_date, end_date)
        rental_result = get_rental_revenue(start_date, end_date)
        rental = rental_result.get('rental', {'sales': 0, 'cogs': 0, 'gross_profit': 0})
        rtr = rental_result.get('rtr', {'sales': 0, 'cogs': 0, 'gross_profit': 0})
        service = get_service_revenue(start_date, end_date)
        parts = get_parts_revenue(start_date, end_date)
        trucking = get_trucking_revenue(start_date, end_date)
        expenses = get_gl_expenses(start_date, end_date)
        ar_aging = get_ar_aging()
        service_calls_per_day = get_service_calls_per_day(start_date, end_date, num_days)
        labor_metrics = get_labor_metrics(start_date, end_date)
        tech_count_data = get_technician_count(start_date, end_date)
        parts_inventory = get_parts_inventory_metrics(start_date, end_date)
        rental_fleet = get_rental_fleet_metrics(start_date, end_date)
        
        # Build data structure for calculate_totals
        data = {
            'new_equipment': new_equipment,
            'rental': rental,
            'rtr': rtr,
            'service': service,
            'parts': parts,
            'trucking': trucking
        }
        
        # Calculate totals
        totals = calculate_totals(data, months_diff)
        
        # Load template
        template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'currie_template.xlsx')
        wb = openpyxl.load_workbook(template_path)
        
        # Delete "New TB" sheet since we're writing actual calculated values now
        if 'New TB' in wb.sheetnames:
            del wb['New TB']
        
        # Get the Sales, COGS, GP sheet
        ws = wb['Sales, COGS, GP']
        
        # Update dealership info - get organization name dynamically
        from flask_jwt_extended import get_jwt_identity
        from src.models.user import User
        user_id = get_jwt_identity()
        user = User.query.get(user_id) if user_id else None
        org_name = user.organization.name if user and user.organization else 'Dealership'
        ws['B3'] = org_name  # Dealership name
        ws['B5'] = datetime.now().strftime('%m/%d/%Y')  # Date
        ws['B7'] = months_diff  # Number of months
        
        # Helper function to write sales/cogs/gp data
        def write_row(row, data):
            ws[f'B{row}'] = data.get('sales', 0)
            ws[f'C{row}'] = data.get('cogs', 0)
            ws[f'D{row}'] = data.get('gross_profit', 0)
        
        # Write Equipment Sales (rows 10-19)
        write_row(10, new_equipment.get('new_lift_truck_primary', {}))
        write_row(11, new_equipment.get('new_lift_truck_other', {}))
        write_row(12, new_equipment.get('new_allied', {}))
        write_row(13, new_equipment.get('other_new_equipment', {}))
        write_row(15, new_equipment.get('operator_training', {}))
        write_row(16, new_equipment.get('used_equipment', {}))
        write_row(17, new_equipment.get('ecommerce', {}))
        write_row(18, new_equipment.get('systems', {}))
        write_row(19, new_equipment.get('batteries', {}))
        
        # Write Rental (row 21), RTR (row 22), Total Rental Dept (row 23)
        write_row(21, rental)
        write_row(22, rtr)  # RTR (Rental Truck Repair) - COGS only, no revenue
        # Total Rental Dept = Rental + RTR
        total_rental_dept = {
            'sales': rental.get('sales', 0) + rtr.get('sales', 0),
            'cogs': rental.get('cogs', 0) + rtr.get('cogs', 0),
            'gross_profit': rental.get('gross_profit', 0) + rtr.get('gross_profit', 0)
        }
        write_row(23, total_rental_dept)
        
        # Write Service (rows 24-28)
        write_row(24, service.get('customer_labor', {}))
        write_row(25, service.get('internal_labor', {}))
        write_row(26, service.get('warranty_labor', {}))
        write_row(27, service.get('sublet', {}))
        write_row(28, service.get('other', {}))  # Fixed: function returns 'other' not 'other_service'
        
        # Write Parts (rows 30-36)
        write_row(30, parts.get('counter_primary', {}))
        write_row(31, parts.get('counter_other', {}))
        write_row(32, parts.get('ro_primary', {}))
        write_row(33, parts.get('ro_other', {}))
        write_row(34, parts.get('internal', {}))  # Fixed: function returns 'internal' not 'internal_parts'
        write_row(35, parts.get('warranty', {}))  # Fixed: function returns 'warranty' not 'warranty_parts'
        write_row(36, parts.get('ecommerce_parts', {}))
        
        # Write Trucking (row 38)
        write_row(38, trucking)
        
        # Get other income/interest data for bottom section
        other_income_interest = get_other_income_and_interest(start_date, end_date)
        
        # Calculate and write bottom section values (rows 40-47)
        # Row 40: Total Aftermarket Sales, COGS & GP
        total_aftermarket = totals['total_aftermarket']
        ws['B40'] = total_aftermarket['sales']
        ws['C40'] = total_aftermarket['cogs']
        ws['D40'] = total_aftermarket['gross_profit']
        
        # Row 41: Total Net Sales & GP
        total_net_sales = totals['total_net_sales_gp']
        ws['B41'] = total_net_sales['sales']
        ws['C41'] = total_net_sales['cogs']
        ws['D41'] = total_net_sales['gross_profit']
        
        # Row 42: Total Company Expenses (write to column B, not D)
        ws['B42'] = expenses['grand_total']
        
        # Row 43: Other Income (Expenses) - negative because it's an expense (write to column B)
        ws['B43'] = other_income_interest.get('other_income', 0)
        
        # Row 44: Interest (Expense) - negative because it's an expense (write to column B)
        ws['B44'] = other_income_interest.get('interest_expense', 0)
        
        # Row 45: Total operating profit = (Total Net Sales GP - Total Expenses + Other Income + Interest) (write to column B)
        total_net_sales_gp = totals['total_net_sales_gp']['gross_profit']
        total_operating_profit = total_net_sales_gp - expenses['grand_total'] + other_income_interest.get('other_income', 0) + other_income_interest.get('interest_expense', 0)
        ws['B45'] = total_operating_profit
        
        # Row 46: F & I Income (write to column B)
        ws['B46'] = other_income_interest.get('fi_income', 0)
        
        # Row 47: Pre-Tax Income = Operating Profit + F&I Income (write to column B)
        ws['B47'] = total_operating_profit + other_income_interest.get('fi_income', 0)
        
        # Write Expenses to "Expenses, Miscellaneous" sheet
        expenses_ws = wb['Expenses, Miscellaneous']
        
        # Department allocation percentages (from tenant-specific Currie mappings)
        currie = get_currie_mappings(schema)
        dept_allocations = currie.get('dept_allocations', {
            'new': 0, 'used': 0, 'rental': 0, 'parts': 0, 'service': 0, 'trucking': 0
        })
        
        # Calculate department expenses for each category
        personnel_total = expenses.get('personnel', {}).get('total', 0)
        operating_total = expenses.get('operating', {}).get('total', 0)
        occupancy_total = expenses.get('occupancy', {}).get('total', 0)
        
        # Allocate Personnel expenses (row 4)
        personnel_new = personnel_total * dept_allocations['new']
        personnel_used = personnel_total * dept_allocations['used']
        personnel_parts = personnel_total * dept_allocations['parts']
        personnel_service = personnel_total * dept_allocations['service']
        personnel_rental = personnel_total * dept_allocations['rental']
        personnel_trucking = personnel_total * dept_allocations['trucking']
        personnel_ga = personnel_total - (personnel_new + personnel_used + personnel_parts + personnel_service + personnel_rental + personnel_trucking)
        
        expenses_ws['B4'] = personnel_new
        expenses_ws['C4'] = personnel_used
        expenses_ws['E4'] = personnel_parts
        expenses_ws['F4'] = personnel_service
        expenses_ws['G4'] = personnel_rental
        expenses_ws['I4'] = personnel_trucking
        expenses_ws['J4'] = personnel_ga
        
        # Allocate Operating expenses (row 5)
        operating_new = operating_total * dept_allocations['new']
        operating_used = operating_total * dept_allocations['used']
        operating_parts = operating_total * dept_allocations['parts']
        operating_service = operating_total * dept_allocations['service']
        operating_rental = operating_total * dept_allocations['rental']
        operating_trucking = operating_total * dept_allocations['trucking']
        operating_ga = operating_total - (operating_new + operating_used + operating_parts + operating_service + operating_rental + operating_trucking)
        
        expenses_ws['B5'] = operating_new
        expenses_ws['C5'] = operating_used
        expenses_ws['E5'] = operating_parts
        expenses_ws['F5'] = operating_service
        expenses_ws['G5'] = operating_rental
        expenses_ws['I5'] = operating_trucking
        expenses_ws['J5'] = operating_ga
        
        # Allocate Occupancy expenses (row 6)
        occupancy_new = occupancy_total * dept_allocations['new']
        occupancy_used = occupancy_total * dept_allocations['used']
        occupancy_parts = occupancy_total * dept_allocations['parts']
        occupancy_service = occupancy_total * dept_allocations['service']
        occupancy_rental = occupancy_total * dept_allocations['rental']
        occupancy_trucking = occupancy_total * dept_allocations['trucking']
        occupancy_ga = occupancy_total - (occupancy_new + occupancy_used + occupancy_parts + occupancy_service + occupancy_rental + occupancy_trucking)
        
        expenses_ws['B6'] = occupancy_new
        expenses_ws['C6'] = occupancy_used
        expenses_ws['E6'] = occupancy_parts
        expenses_ws['F6'] = occupancy_service
        expenses_ws['G6'] = occupancy_rental
        expenses_ws['I6'] = occupancy_trucking
        expenses_ws['J6'] = occupancy_ga
        # Row 7 has formulas =SUM(B4:B6) etc. which will calculate automatically
        
        # AR Aging (rows 22-26, column B)
        expenses_ws['B22'] = ar_aging.get('current', 0)  # Current
        expenses_ws['B23'] = ar_aging.get('days_31_60', 0)  # 31-60 days
        expenses_ws['B24'] = ar_aging.get('days_61_90', 0)  # 61-90 days
        expenses_ws['B25'] = ar_aging.get('days_91_plus', 0)  # 91+ days
        # B26 has a formula =SUM(B22:B25) which will calculate automatically
        
        # Parts Department Metrics (column E, rows 22-26)
        expenses_ws['E22'] = parts_inventory.get('fill_rate', 0) / 100 if parts_inventory.get('fill_rate', 0) else 0  # Stock Fill Rate as decimal
        # E23 (Stock v. NonStock) - not available from Softbase
        # E24 (Parts Inventory Aging %) - percentage of inventory over 12 months
        if parts_inventory.get('aging'):
            aging = parts_inventory['aging']
            total_parts = aging.get('obsolete_count', 0) + aging.get('slow_count', 0) + aging.get('medium_count', 0) + aging.get('fast_count', 0)
            if total_parts > 0:
                expenses_ws['E24'] = aging.get('obsolete_count', 0) / total_parts  # % over 12 months
        
        # Technician Productivity (column H, rows 22-28)
        tech_count = tech_count_data.get('active_technicians', 0)
        total_billed_hours = labor_metrics.get('total_billed_hours', 0)
        avg_labor_rate = labor_metrics.get('average_labor_rate', 0)
        
        # H22: # of units under full maintenance contract - not directly available
        expenses_ws['G23'] = round(avg_labor_rate, 2) if avg_labor_rate else None  # Customer Labor Rate (G23:H23 merged)
        # H24: Avg. Hourly Tech Pay Rate - not available from Softbase GL
        expenses_ws['H25'] = round(total_billed_hours, 1) if total_billed_hours else None  # Total Hours Billed
        # H26: Productive Hours - need to split billed vs non-productive (not available)
        # H27: Non-Productive Hours - not available
        # H28: Total Hours Paid - not available from Softbase
        
        # Additional Technician Productivity (column L, rows 28-30)
        # L28: PM Completion Rate - not available
        # L29: First Call Completion Rate - not available
        # L30: Average Response Time - not available
        
        # Marketshare Information (column E, rows 29-31)
        # E29: Sold Units - requires external data, not available from Softbase
        # E30: Lost Units - requires external data
        # E31: Size of Market - requires external data
        
        # ST Rental Fleet Metrics (column E-F, rows 34-40)
        rental_unit_count = rental_fleet.get('unit_count', 0)
        rental_gross_value = rental_fleet.get('gross_fleet_value', 0)
        rental_depreciation = rental_fleet.get('annualized_depreciation', 0)
        
        # Scale depreciation/interest to match the reporting period (not annualized)
        period_scale = num_days / 365 if num_days > 0 else 1
        period_depreciation = rental_depreciation * period_scale
        
        expenses_ws['D34'] = round(period_depreciation, 2) if period_depreciation else None  # ST Rental Depreciation (D34:E34 merged)
        # E35: Interest - would need separate GL query for rental interest expense
        # E36: Maintenance - would need separate GL query for rental maintenance expense
        expenses_ws['F38'] = rental_unit_count if rental_unit_count > 0 else None  # # of Units in ST Rental Fleet (was hardcoded 440)
        expenses_ws['D39'] = round(rental_gross_value, 2) if rental_gross_value else None  # Acquisition Cost of ST Fleet (D39:E39 merged)
        
        # Service Calls per day (cell L37)
        expenses_ws['L37'] = service_calls_per_day.get('calls_per_day', 0)
        
        # Write Balance Sheet data
        balance_sheet_data = get_balance_sheet_data(end_date)
        bs_ws = wb['Balance Sheet']
        
        # Helper function to sum account balances
        def sum_accounts(account_list):
            return sum(acc.get('balance', 0) for acc in account_list)
        
        # ASSETS
        assets = balance_sheet_data['assets']
        
        # Cash (B6) - sum all cash accounts
        bs_ws['B6'] = sum_accounts(assets['current_assets']['cash'])
        
        # Trade Accounts Receivable (B7) - sum AR accounts
        bs_ws['B7'] = sum_accounts(assets['current_assets']['accounts_receivable'])
        
        # All Other Accounts Receivable (B8) - set to 0 or use other current assets if needed
        bs_ws['B8'] = 0
        
        # Inventory breakdown - use tenant-specific description patterns
        inventory_accounts = assets['current_assets']['inventory']
        inv_patterns = currie.get('inventory_patterns', {})
        
        # Helper: check if description matches any pattern in a list
        def desc_matches(description, patterns):
            desc_upper = description.upper()
            return any(p in desc_upper for p in patterns)
        
        new_equipment_primary = 0
        new_equipment_other = 0
        new_allied_inventory = 0
        other_new_equipment = 0
        used_equipment_inventory = 0
        parts_inventory = 0
        battery_inventory = 0
        wip_balance = 0
        other_inventory = 0
        
        for acc in inventory_accounts:
            desc = acc['description']
            balance = acc['balance']
            if desc_matches(desc, inv_patterns.get('wip', [])):
                wip_balance += balance
            elif desc_matches(desc, inv_patterns.get('new_equipment_primary', [])):
                new_equipment_primary += balance
            elif desc_matches(desc, inv_patterns.get('new_allied_inventory', [])):
                new_allied_inventory += balance
            elif desc_matches(desc, inv_patterns.get('used_equipment_inventory', [])):
                used_equipment_inventory += balance
            elif desc_matches(desc, inv_patterns.get('parts_inventory', [])):
                parts_inventory += balance
            elif desc_matches(desc, inv_patterns.get('battery_inventory', [])):
                battery_inventory += balance
            else:
                other_inventory += balance
        
        bs_ws['B11'] = new_equipment_primary  # New Equipment, primary brand
        bs_ws['B12'] = new_equipment_other  # New Equipment, other brand (currently $0)
        bs_ws['B13'] = new_allied_inventory  # New Allied Inventory
        bs_ws['B14'] = other_new_equipment  # Other New Equipment (currently $0)
        bs_ws['B15'] = used_equipment_inventory  # Used Equipment Inventory
        bs_ws['B16'] = parts_inventory  # Parts Inventory
        bs_ws['B17'] = battery_inventory  # Battery Inventory
        bs_ws['B18'] = other_inventory  # Other Inventory
        bs_ws['B21'] = wip_balance  # Work in Progress
        
        # Other Current Assets (B23)
        bs_ws['B23'] = sum_accounts(assets['current_assets']['other_current'])
        
        # Fixed Assets - use tenant-specific patterns for rental fleet identification
        fa_patterns = currie.get('fixed_asset_patterns', {})
        rental_fleet_patterns = fa_patterns.get('rental_fleet', [])
        rental_fleet_net = 0
        other_fixed = 0
        
        for acc in assets['fixed_assets']:
            desc = acc['description']
            balance = acc['balance']
            if desc_matches(desc, rental_fleet_patterns):
                rental_fleet_net += balance
            else:
                other_fixed += balance
        
        bs_ws['B27'] = rental_fleet_net  # Rental Fleet (net of depreciation)
        bs_ws['B28'] = other_fixed  # Other Long Term or Fixed Assets
        
        # Other Assets (B29)
        bs_ws['B29'] = sum_accounts(assets['other_assets'])
        
        # LIABILITIES - match Currie web page structure exactly
        liabilities = balance_sheet_data['liabilities']
        
        # Current Liabilities breakdown - use tenant-specific patterns
        liab_patterns = currie.get('liability_patterns', {})
        cl_patterns = liab_patterns.get('current', {})
        lt_patterns = liab_patterns.get('long_term', {})
        
        ap_primary = 0
        ap_other = 0
        notes_payable_current = 0
        short_term_rental_finance = 0
        used_equipment_financing = 0
        other_current_liabilities = 0
        
        for acc in liabilities['current_liabilities']:
            desc = acc['description']
            balance = acc['balance']
            if desc_matches(desc, cl_patterns.get('ap_primary', [])):
                ap_primary += balance
            elif desc_matches(desc, cl_patterns.get('short_term_rental_finance', [])):
                short_term_rental_finance += balance
            elif desc_matches(desc, cl_patterns.get('used_equipment_financing', [])):
                used_equipment_financing += balance
            else:
                other_current_liabilities += balance
        
        # Negate liability values for standard BS presentation (credit balances shown as positive)
        bs_ws['E7'] = -ap_primary  # A/P Primary Brand
        bs_ws['E8'] = -ap_other  # A/P Other
        bs_ws['E9'] = -notes_payable_current  # Notes Payable - due within 1 year
        bs_ws['E10'] = -short_term_rental_finance  # Short Term Rental Finance
        bs_ws['E11'] = -used_equipment_financing  # Used Equipment Financing
        bs_ws['E12'] = -other_current_liabilities  # Other Current Liabilities
        
        # Long-term Liabilities breakdown - use tenant-specific patterns
        long_term_notes = 0
        loans_from_stockholders = 0
        lt_rental_fleet_financing = 0
        
        for acc in liabilities['long_term_liabilities']:
            desc = acc['description']
            balance = acc['balance']
            # Check patterns in priority order: floorplan first (most specific),
            # then stockholders, then general notes payable (least specific)
            if desc_matches(desc, lt_patterns.get('lt_rental_fleet_financing', [])):
                lt_rental_fleet_financing += balance
            elif desc_matches(desc, lt_patterns.get('loans_from_stockholders', [])):
                loans_from_stockholders += balance
            elif desc_matches(desc, lt_patterns.get('long_term_notes', [])):
                long_term_notes += balance
        
        total_long_term_liabilities = sum(acc['balance'] for acc in liabilities['long_term_liabilities'])
        other_long_term_debt = total_long_term_liabilities - (long_term_notes + loans_from_stockholders + lt_rental_fleet_financing)
        
        # Other Liabilities
        other_liab_remaining = sum_accounts(liabilities['other_liabilities'])
        
        bs_ws['E15'] = -long_term_notes  # Long Term Notes Payable
        bs_ws['E16'] = -loans_from_stockholders  # Loans from Stockholders
        bs_ws['E17'] = -lt_rental_fleet_financing  # LT Rental Fleet Financing
        bs_ws['E18'] = -other_long_term_debt  # Other Long Term Debt
        
        # Other Liabilities (E23)
        bs_ws['E23'] = -other_liab_remaining
        
        # EQUITY
        equity = balance_sheet_data['equity']
        
        # Negate equity values for standard BS presentation (credit balances shown as positive)
        # Capital Stock (E27)
        capital_stock_total = -sum_accounts(equity['capital_stock'])
        bs_ws['E27'] = capital_stock_total
        
        # Retained Earnings (E28) - includes distributions
        retained_earnings_total = -(sum_accounts(equity['retained_earnings']) + sum_accounts(equity['distributions']))
        bs_ws['E28'] = retained_earnings_total
        
        # Total Net Worth (E29) = Capital Stock + Retained Earnings + Current Year Net Income
        net_income = -equity.get('net_income', 0)  # Negate: GL net income sign is opposite
        bs_ws['E29'] = capital_stock_total + retained_earnings_total + net_income
        
        # Balance check formula (E31): Both sides now positive, use tolerance for rounding
        from openpyxl.styles import Font
        bs_ws['E31'] = '=IF(ABS(E30-B30)<1,"Balanced","Not Balanced")'
        bs_ws['E31'].font = Font(bold=True, color='00008000')  # Green for balanced
        
        # Save to BytesIO for download
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with date range
        filename = f"Currie_Financial_Model_{start_date}_to_{end_date}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting Currie Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to export Excel', 'message': str(e)}), 500


@currie_bp.route('/api/currie/discover-gl-accounts', methods=['GET'])
@jwt_required()
def discover_gl_accounts():
    """Debug endpoint to discover GL expense accounts"""
    try:
        # Get date range from query parameters
        start_date = request.args.get('start_date', '2025-09-01')
        end_date = request.args.get('end_date', '2025-11-30')
        
        schema = get_tenant_schema()

        
        query = f"""
        SELECT 
            AccountNo,
            COUNT(*) as TransactionCount,
            SUM(Amount) as TotalAmount
        FROM {schema}.GLDetail
        WHERE Posted = 1
          AND EffectiveDate >= %s
          AND EffectiveDate <= %s
          AND (
            AccountNo LIKE '5%' OR 
            AccountNo LIKE '6%' OR 
            AccountNo LIKE '7%'
          )
        GROUP BY AccountNo
        ORDER BY AccountNo
        """
        
        results = get_sql_service().execute_query(query, [start_date, end_date])
        
        # Format for easy reading
        accounts = []
        for r in results:
            accounts.append({
                'account_no': r['AccountNo'],
                'transaction_count': int(r['TransactionCount']),
                'total_amount': float(r['TotalAmount'] or 0)
            })
        
        return jsonify({
            'date_range': {
                'start': start_date,
                'end': end_date
            },
            'total_accounts': len(accounts),
            'accounts': accounts
        })
        
    except Exception as e:
        logger.error(f"Error discovering GL accounts: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def is_full_calendar_month(start_date, end_date):
    """
    Check if the date range represents a full calendar month
    Excludes the current month to ensure GL.MTD is only used for closed/completed months
    
    Args:
        start_date: Start date string (YYYY-MM-DD)
        end_date: End date string (YYYY-MM-DD)
    
    Returns:
        Tuple of (is_full_month, year, month) or (False, None, None)
    """
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        now = datetime.now()
        
        # Check if start is first day of month
        if start.day != 1:
            return False, None, None
        
        # Check if end is last day of month
        last_day = calendar.monthrange(start.year, start.month)[1]
        if end.day != last_day:
            return False, None, None
        
        # Check if both dates are in the same month
        if start.year != end.year or start.month != end.month:
            return False, None, None
        
        # Exclude current month - always use GLDetail for in-process months
        if start.year == now.year and start.month == now.month:
            return False, None, None
        
        return True, start.year, start.month
    except:
        return False, None, None

def get_gl_expenses(start_date, end_date):
    """
    Get operating expenses from GL accounts
    Uses GL.MTD for full calendar months (exact Softbase match)
    Uses GLDetail for custom date ranges (flexibility)
    """
    try:
        # Check if this is a full calendar month
        is_full_month, year, month = is_full_calendar_month(start_date, end_date)
        
        if is_full_month:
            # Use GL.MTD for exact Softbase match
            return get_gl_expenses_from_gl_mtd(year, month)
        else:
            # Use GLDetail for custom date ranges
            return get_gl_expenses_from_gldetail(start_date, end_date)
    except Exception as e:
        logger.error(f"Error in get_gl_expenses: {e}")
        raise

def _build_expense_category_query(schema, table, amount_col, category_config, where_clause, params):
    """
    Build a dynamic expense query from currie mappings.
    category_config: {'accounts': [...], 'detail': {'key': [accts], ...}}
    Returns: (result_dict, total)
    """
    all_accounts = category_config.get('accounts', [])
    detail = category_config.get('detail', {})
    
    if not all_accounts:
        result = {k: 0 for k in detail.keys()}
        result['total'] = 0
        return result
    
    # Build CASE WHEN fragments for each detail key
    case_fragments = []
    for key, accts in detail.items():
        if not accts:
            case_fragments.append(f"0 as {key}")
        elif len(accts) == 1:
            case_fragments.append(f"SUM(CASE WHEN AccountNo = '{accts[0]}' THEN {amount_col} ELSE 0 END) as {key}")
        else:
            quoted = ','.join([f"'{a}'" for a in accts])
            case_fragments.append(f"SUM(CASE WHEN AccountNo IN ({quoted}) THEN {amount_col} ELSE 0 END) as {key}")
    
    # Add total
    all_quoted = ','.join([f"'{a}'" for a in all_accounts])
    case_fragments.insert(0, f"SUM({amount_col}) as category_total")
    
    select_clause = ',\n            '.join(case_fragments)
    
    query = f"""
    SELECT 
        {select_clause}
    FROM {schema}.{table}
    WHERE {where_clause}
      AND AccountNo IN ({all_quoted})
    """
    
    query_result = get_sql_service().execute_query(query, params)
    row = query_result[0] if query_result else {}
    
    result = {}
    for key in detail.keys():
        result[key] = float(row.get(key) or 0)
    result['total'] = float(row.get('category_total') or 0)
    
    return result


def _build_expenses_response(schema, table, amount_col, where_clause, params):
    """
    Build the full expenses response using currie mappings.
    Shared by both gl_mtd and gldetail expense functions.
    """
    currie = get_currie_mappings(schema)
    expense_config = currie.get('expenses', {})
    
    personnel = _build_expense_category_query(
        schema, table, amount_col,
        expense_config.get('personnel', {'accounts': [], 'detail': {}}),
        where_clause, params
    )
    occupancy = _build_expense_category_query(
        schema, table, amount_col,
        expense_config.get('occupancy', {'accounts': [], 'detail': {}}),
        where_clause, params
    )
    operating = _build_expense_category_query(
        schema, table, amount_col,
        expense_config.get('operating', {'accounts': [], 'detail': {}}),
        where_clause, params
    )
    
    return {
        'personnel': personnel,
        'occupancy': occupancy,
        'operating': operating,
        'grand_total': personnel['total'] + occupancy['total'] + operating['total']
    }


def get_gl_expenses_from_gl_mtd(year, month):
    """
    Get operating expenses from GL.MTD (monthly summary table)
    This matches Softbase exactly for monthly reports
    """
    try:
        schema = get_tenant_schema()
        return _build_expenses_response(
            schema, 'GL', 'MTD',
            'Year = %s AND Month = %s',
            [year, month]
        )
    except Exception as e:
        logger.error(f"Error fetching GL expenses from GL.MTD: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'personnel': {'total': 0},
            'occupancy': {'total': 0},
            'operating': {'total': 0},
            'grand_total': 0
        }

def get_gl_expenses_from_gldetail(start_date, end_date):
    """
    Get operating expenses from GLDetail (transaction-level detail)
    Used for custom date ranges that aren't full calendar months
    """
    try:
        schema = get_tenant_schema()
        return _build_expenses_response(
            schema, 'GLDetail', 'Amount',
            'Posted = 1 AND EffectiveDate >= %s AND EffectiveDate <= %s',
            [start_date, end_date]
        )
    except Exception as e:
        logger.error(f"Error fetching GL expenses: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'personnel': {'total': 0},
            'occupancy': {'total': 0},
            'operating': {'total': 0},
            'grand_total': 0
        }


@currie_bp.route('/api/currie/expenses', methods=['GET'])
@jwt_required()
def get_currie_expenses():
    """Get operating expenses for Currie Financial Model"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({'error': 'start_date and end_date are required'}), 400
        
        expenses = get_gl_expenses(start_date, end_date)
        
        return jsonify(expenses)
        
    except Exception as e:
        logger.error(f"Error in get_currie_expenses: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def get_other_income_and_interest(start_date, end_date):
    """
    Get Other Income (Expenses), Interest (Expense), and F&I Income
    These appear in the bottom summary section of the Currie report
    Account mappings are loaded from tenant-specific currie mappings.
    """
    try:
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        oi_config = currie.get('other_income_interest', {})
        
        other_exp_accts = oi_config.get('other_expenses', [])
        interest_accts = oi_config.get('interest_expense', [])
        fi_accts = oi_config.get('fi_income', [])
        
        all_accounts = other_exp_accts + interest_accts + fi_accts
        if not all_accounts:
            return {'other_income': 0, 'interest_expense': 0, 'fi_income': 0}
        
        # Build dynamic CASE WHEN
        other_quoted = ','.join([f"'{a}'" for a in other_exp_accts]) if other_exp_accts else "'NONE'"
        interest_quoted = ','.join([f"'{a}'" for a in interest_accts]) if interest_accts else "'NONE'"
        fi_quoted = ','.join([f"'{a}'" for a in fi_accts]) if fi_accts else "'NONE'"
        all_quoted = ','.join([f"'{a}'" for a in all_accounts])

        query = f"""
        SELECT 
            SUM(CASE WHEN AccountNo IN ({other_quoted}) THEN Amount ELSE 0 END) as other_expenses,
            SUM(CASE WHEN AccountNo IN ({interest_quoted}) THEN Amount ELSE 0 END) as interest_expense,
            SUM(CASE WHEN AccountNo IN ({fi_quoted}) THEN Amount ELSE 0 END) as fi_income
        FROM {schema}.GLDetail
        WHERE EffectiveDate >= %s 
          AND EffectiveDate <= %s
          AND Posted = 1
          AND AccountNo IN ({all_quoted})
        """
        
        result = get_sql_service().execute_query(query, [start_date, end_date])
        
        if result and len(result) > 0:
            row = result[0]
            # Return as negative because these are expenses that reduce operating profit
            # Excel formulas use -SUM() to make them negative
            return {
                'other_income': -float(row.get('other_expenses') or 0),  # Negative because it's an expense
                'interest_expense': -float(row.get('interest_expense') or 0),  # Negative because it's an expense
                'fi_income': float(row.get('fi_income') or 0)  # Positive because it's income
            }
        else:
            return {
                'other_income': 0,
                'interest_expense': 0,
                'fi_income': 0
            }
            
    except Exception as e:
        logger.error(f"Error fetching other income and interest: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'other_income': 0,
            'interest_expense': 0,
            'fi_income': 0
        }


def get_balance_sheet_data(as_of_date):
    """
    Get Balance Sheet data from GL accounts.
    Uses tenant-specific balance_sheet_categories from currie_mappings for account prefix routing.
    Assets: 1xxxxx (Bennett) or 1xxxxxx (IPS)
    Liabilities: 2xxxxx or 2xxxxxx
    Equity: 3xxxxx or 3xxxxxx
    
    Returns categorized balance sheet accounts with balances as of the specified date
    """
    try:
        # Parse the date to get year and month
        date_obj = datetime.strptime(as_of_date, '%Y-%m-%d')
        year = date_obj.year
        month = date_obj.month
        
        # Get tenant-specific balance sheet categorization rules
        schema = get_tenant_schema()
        currie = get_currie_mappings(schema)
        bs_cats = currie.get('balance_sheet_categories', {})
        asset_cats = bs_cats.get('assets', {})
        liab_cats = bs_cats.get('liabilities', {})
        equity_cats = bs_cats.get('equity', {})
        
        # Helper: check if account_no starts with any prefix in a list
        def matches_prefixes(account_no, prefixes):
            return any(account_no.startswith(p) for p in prefixes)

        # Query GL.YTD for balance sheet accounts
        query = f"""
        SELECT 
            gl.AccountNo,
            COALESCE(coa.Description, gl.AccountNo) as Description,
            gl.YTD as balance
        FROM {schema}.GL gl
        LEFT JOIN {schema}.ChartOfAccounts coa ON gl.AccountNo = coa.AccountNo
        WHERE gl.Year = %s 
          AND gl.Month = %s
          AND (
            gl.AccountNo LIKE '1%'  -- Assets
            OR gl.AccountNo LIKE '2%'  -- Liabilities
            OR gl.AccountNo LIKE '3%'  -- Equity
          )
          AND gl.YTD != 0
        ORDER BY gl.AccountNo
        """
        
        result = get_sql_service().execute_query(query, [year, month])
        
        logger.info(f"Balance Sheet query for year={year}, month={month}, schema={schema}")
        logger.info(f"Query returned {len(result) if result else 0} rows")
        if result:
            logger.info(f"First row sample: {result[0]}")
        
        # Calculate current year net income from ALL P&L accounts
        net_income_query = f"""
        SELECT COALESCE(SUM(gl.YTD), 0) as net_income
        FROM {schema}.GL gl
        WHERE gl.Year = %s 
          AND gl.Month = %s
          AND gl.AccountNo NOT LIKE '1%'
          AND gl.AccountNo NOT LIKE '2%'
          AND gl.AccountNo NOT LIKE '3%'
        """
        net_income_result = get_sql_service().execute_query(net_income_query, [year, month])
        net_income = float(net_income_result[0].get('net_income', 0)) if net_income_result else 0
        logger.info(f"Current year net income (P&L accounts): {net_income:,.2f}")
        
        # Initialize categorized structures
        assets = {
            'current_assets': {
                'cash': [],
                'accounts_receivable': [],
                'inventory': [],
                'other_current': []
            },
            'fixed_assets': [],
            'other_assets': [],
            'total': 0
        }
        
        liabilities = {
            'current_liabilities': [],
            'long_term_liabilities': [],
            'other_liabilities': [],
            'total': 0
        }
        
        equity = {
            'capital_stock': [],
            'retained_earnings': [],
            'distributions': [],
            'net_income': net_income,
            'total': 0
        }
        
        if result:
            for row in result:
                account_no = str(row.get('AccountNo', '')).strip()
                description = row.get('Description', '').strip()
                balance = float(row.get('balance', 0))
                
                account_data = {
                    'account': account_no,
                    'description': description,
                    'balance': balance
                }
                
                # --- ASSETS ---
                if account_no.startswith('1'):
                    categorized = False
                    # Check each asset sub-category using tenant-specific prefixes
                    if matches_prefixes(account_no, asset_cats.get('cash', [])):
                        assets['current_assets']['cash'].append(account_data)
                        categorized = True
                    elif matches_prefixes(account_no, asset_cats.get('accounts_receivable', [])):
                        assets['current_assets']['accounts_receivable'].append(account_data)
                        categorized = True
                    elif matches_prefixes(account_no, asset_cats.get('inventory', [])):
                        assets['current_assets']['inventory'].append(account_data)
                        categorized = True
                    elif matches_prefixes(account_no, asset_cats.get('other_current', [])):
                        assets['current_assets']['other_current'].append(account_data)
                        categorized = True
                    elif matches_prefixes(account_no, asset_cats.get('fixed_assets', [])):
                        assets['fixed_assets'].append(account_data)
                        categorized = True
                    
                    if not categorized:
                        assets['other_assets'].append(account_data)
                    
                    assets['total'] += balance
                
                # --- LIABILITIES ---
                elif account_no.startswith('2'):
                    if matches_prefixes(account_no, liab_cats.get('current', [])):
                        liabilities['current_liabilities'].append(account_data)
                    elif matches_prefixes(account_no, liab_cats.get('long_term', [])):
                        liabilities['long_term_liabilities'].append(account_data)
                    else:
                        liabilities['other_liabilities'].append(account_data)
                    
                    liabilities['total'] += balance
                
                # --- EQUITY ---
                elif account_no.startswith('3'):
                    if matches_prefixes(account_no, equity_cats.get('capital_stock', [])):
                        equity['capital_stock'].append(account_data)
                    elif matches_prefixes(account_no, equity_cats.get('distributions', [])):
                        equity['distributions'].append(account_data)
                    elif matches_prefixes(account_no, equity_cats.get('retained_earnings', [])):
                        equity['retained_earnings'].append(account_data)
                    else:
                        equity['retained_earnings'].append(account_data)  # Default to retained earnings
                    
                    equity['total'] += balance
        
        # Add current year net income to equity total
        equity['total'] += net_income
        
        return {
            'assets': assets,
            'liabilities': liabilities,
            'equity': equity,
            'as_of_date': as_of_date,
            'balanced': abs(assets['total'] + liabilities['total'] + equity['total']) < 0.01
        }
        
    except Exception as e:
        logger.error(f"Error fetching balance sheet data: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'assets': {'current_assets': {'cash': [], 'accounts_receivable': [], 'inventory': [], 'other_current': []}, 'fixed_assets': [], 'other_assets': [], 'total': 0},
            'liabilities': {'current_liabilities': [], 'long_term_liabilities': [], 'other_liabilities': [], 'total': 0},
            'equity': {'capital_stock': [], 'retained_earnings': [], 'distributions': [], 'total': 0},
            'as_of_date': as_of_date,
            'balanced': False,
            'error': str(e)
        }
