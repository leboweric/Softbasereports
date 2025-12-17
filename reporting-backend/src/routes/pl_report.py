"""
P&L (Profit & Loss) Report Route
Generates departmental and consolidated P&L reports using GLDetail with exact GL account mappings
"""

from flask import Blueprint, jsonify, request
from datetime import datetime
import logging
import calendar
from src.services.azure_sql_service import AzureSQLService
from src.utils.fiscal_year import get_fiscal_ytd_start

logger = logging.getLogger(__name__)
pl_report_bp = Blueprint('pl_report', __name__)
sql_service = AzureSQLService()

# GL Account Mappings by Department (Source: Softbase P&L)
GL_ACCOUNTS = {
    'new_equipment': {
        'dept_code': 10,
        'dept_name': 'New Equipment',
        'revenue': ['410001', '412001', '413001', '414001', '421001', '426001', '431001', '434001'],
        'cogs': ['510001', '513001', '514001', '521001', '525001', '526001', '531001', '534001', '534013', '538000']
    },
    'used_equipment': {
        'dept_code': 20,
        'dept_name': 'Used Equipment',
        'revenue': ['410002', '412002', '413002', '414002', '421002', '426002', '431002', '434002', '436001'],
        'cogs': ['510002', '512002', '513002', '514002', '521002', '525002', '526002', '531002', '534002', '536001']
    },
    'parts': {
        'dept_code': 30,
        'dept_name': 'Parts',
        'revenue': ['410003', '410012', '410014', '410015', '421003', '424000', '429001', '430000', '433000', '434003', '436002', '439000'],
        'cogs': ['510003', '510012', '510013', '510014', '510015', '521003', '522001', '524000', '529002', '530000', '533000', '534003', '536002', '542000', '543000', '544000']
    },
    'service': {
        'dept_code': 40,
        'dept_name': 'Service',
        'revenue': ['410004', '410005', '410007', '410016', '421004', '421005', '421006', '421007', '423000', '425000', '428000', '429002', '432000', '435000', '435001', '435002', '435003', '435004'],
        'cogs': ['510004', '510005', '510007', '512001', '521004', '521005', '521006', '521007', '522000', '523000', '528000', '529001', '534015', '535001', '535002', '535003', '535004', '535005']
    },
    'rental': {
        'dept_code': 60,
        'dept_name': 'Rental',
        'revenue': ['410008', '411001', '419000', '420000', '421000', '434012'],
        'cogs': ['510008', '511001', '519000', '520000', '521008', '534014', '537001', '539000', '545000']
    },
    'transportation': {
        'dept_code': 80,
        'dept_name': 'Transportation',
        'revenue': ['410010', '421010', '434010', '434013'],
        'cogs': ['510010', '521010', '534010', '534012']
    },
    'administrative': {
        'dept_code': 90,
        'dept_name': 'Administrative',
        'revenue': ['410011', '421011', '422100', '427000', '434011'],
        'cogs': ['510011', '521011', '522100', '525000', '527000', '532000', '534011', '540000', '541000']
    }
}

# Other Income/Contra-Revenue Accounts (7xxxxx series)
# These appear in Softbase under "Sales" category but are other income/adjustments
# Note: 706000 (ADMINISTRATIVE FUND EXPENSE) is NOT included - it's an expense account
OTHER_INCOME_ACCOUNTS = ['701000', '702000', '703000', '704000', '705000']

# Expense Account Mappings (all in Administrative department)
# Including all 6xxxxx accounts (some may have $0 in CSV but non-zero in GLDetail)
EXPENSE_ACCOUNTS = {
    'depreciation': ['600900'],
    'salaries_wages': ['602000', '602001', '602300', '602301', '602302', '602600', '602610'],
    'payroll_benefits': ['601100', '602700', '602701'],
    'rent_facilities': ['600200', '600201', '600300', '602100'],
    'utilities': ['604000'],
    'insurance': ['601700'],
    'marketing': ['600000', '603300'],
    'professional_fees': ['603000'],
    'office_admin': ['600500', '601300', '602400', '602900', '603500', '603600'],
    'vehicle_equipment': ['604100'],
    'interest_finance': ['601800', '602500'],
    'other_expenses': [
        '600100', '600400', '600600', '600700', '600800', '600901', '600902', '601000', '601200', 
        '601400', '601500', '601600', '601900', '602200', '602601', '602800', 
        '603100', '603101', '603102', '603103', '603200', '603400', '603501', 
        '603700', '603800', '603900', '604200', '650000', '706000', '999999'
    ]
}
def get_department_data(start_date, end_date, dept_key, include_detail=False):
    """
    Get P&L data for a specific department
    Uses GL.MTD for full calendar months (exact Softbase match)
    Uses GLDetail for custom date ranges (flexibility)
    
    Args:
        start_date: Start date for the report
        end_date: End date for the report
        dept_key: Department key from GL_ACCOUNTS
        include_detail: If True, include account-level detail
    
    Returns:
        Dictionary with revenue, cogs, gross_profit, gross_margin, and optionally account details
    """
    try:
        # Check if this is a full calendar month
        is_full_month, year, month = is_full_calendar_month(start_date, end_date)
        
        if is_full_month:
            # Use GL.MTD for exact Softbase match
            return get_department_data_from_gl_mtd(year, month, dept_key, include_detail)
        else:
            # Use GLDetail for custom date ranges
            return get_department_data_from_gldetail(start_date, end_date, dept_key, include_detail)
    except Exception as e:
        logger.error(f"Error in get_department_data for {dept_key}: {e}")
        raise

def get_department_data_from_gl_mtd(year, month, dept_key, include_detail=False):
    """
    Get department P&L data from GL.MTD (monthly summary table)
    This matches Softbase exactly for monthly reports
    """
    try:
        dept_config = GL_ACCOUNTS[dept_key]
        revenue_accounts = dept_config['revenue']
        cogs_accounts = dept_config['cogs']
        
        # Build account lists for SQL IN clause
        all_accounts = revenue_accounts + cogs_accounts
        account_list = "', '".join(all_accounts)
        
        if include_detail:
            # Get account-level detail from GL.MTD
            query = f"""
            SELECT 
                AccountNo,
                -MTD as total
            FROM ben002.GL
            WHERE Year = %s
              AND Month = %s
              AND AccountNo IN ('{account_list}')
            """
            
            results = sql_service.execute_query(query, [year, month])
            
            revenue = 0
            cogs = 0
            revenue_detail = []
            cogs_detail = []
            
            for row in results:
                account_no = row['AccountNo']
                total = float(row['total'] or 0)
                
                if account_no in revenue_accounts:
                    revenue += total
                    revenue_detail.append({'account': account_no, 'amount': total})
                elif account_no in cogs_accounts:
                    cogs += total
                    cogs_detail.append({'account': account_no, 'amount': total})
            
            gross_profit = revenue - cogs
            gross_margin = (gross_profit / revenue * 100) if revenue > 0 else 0
            
            return {
                'dept_code': dept_config['dept_code'],
                'dept_name': dept_config['dept_name'],
                'revenue': revenue,
                'cogs': cogs,
                'gross_profit': gross_profit,
                'gross_margin': gross_margin,
                'revenue_detail': revenue_detail,
                'cogs_detail': cogs_detail
            }
        else:
            # Get summary only from GL.MTD
            revenue_list = "', '".join(revenue_accounts)
            cogs_list = "', '".join(cogs_accounts)
            
            query = f"""
            SELECT 
                -SUM(CASE WHEN AccountNo IN ('{revenue_list}') THEN MTD ELSE 0 END) as revenue,
                SUM(CASE WHEN AccountNo IN ('{cogs_list}') THEN MTD ELSE 0 END) as cogs
            FROM ben002.GL
            WHERE Year = %s
              AND Month = %s
              AND AccountNo IN ('{account_list}')
            """
            
            results = sql_service.execute_query(query, [year, month])
            
            if results and len(results) > 0:
                row = results[0]
                revenue = float(row['revenue'] or 0)
                cogs = float(row['cogs'] or 0)
                gross_profit = revenue - cogs
                gross_margin = (gross_profit / revenue * 100) if revenue > 0 else 0
                
                return {
                    'dept_code': dept_config['dept_code'],
                    'dept_name': dept_config['dept_name'],
                    'revenue': revenue,
                    'cogs': cogs,
                    'gross_profit': gross_profit,
                    'gross_margin': gross_margin
                }
        
        return {
            'dept_code': dept_config['dept_code'],
            'dept_name': dept_config['dept_name'],
            'revenue': 0,
            'cogs': 0,
            'gross_profit': 0,
            'gross_margin': 0
        }
        
    except Exception as e:
        logger.error(f"Error fetching P&L from GL.MTD for {dept_key}: {str(e)}")
        raise

def get_department_data_from_gldetail(start_date, end_date, dept_key, include_detail=False):
    """
    Get department P&L data from GLDetail (transaction-level detail)
    Used for custom date ranges that aren't full calendar months
    """
    try:
        dept_config = GL_ACCOUNTS[dept_key]
        revenue_accounts = dept_config['revenue']
        cogs_accounts = dept_config['cogs']
        
        # Build account lists for SQL IN clause
        all_accounts = revenue_accounts + cogs_accounts
        account_list = "', '".join(all_accounts)
        
        if include_detail:
            # Get account-level detail
            query = f"""
            SELECT 
                AccountNo,
                -SUM(Amount) as total
            FROM ben002.GLDetail
            WHERE EffectiveDate >= %s 
              AND EffectiveDate <= %s
              AND Posted = 1
              AND AccountNo IN ('{account_list}')
            GROUP BY AccountNo
            """
            
            results = sql_service.execute_query(query, [start_date, end_date])
            
            revenue = 0
            cogs = 0
            revenue_detail = []
            cogs_detail = []
            
            for row in results:
                account_no = row['AccountNo']
                total = float(row['total'] or 0)
                
                if account_no in revenue_accounts:
                    revenue += total
                    revenue_detail.append({'account': account_no, 'amount': total})
                elif account_no in cogs_accounts:
                    cogs += total
                    cogs_detail.append({'account': account_no, 'amount': total})
            
            gross_profit = revenue - cogs
            gross_margin = (gross_profit / revenue * 100) if revenue > 0 else 0
            
            return {
                'dept_code': dept_config['dept_code'],
                'dept_name': dept_config['dept_name'],
                'revenue': revenue,
                'cogs': cogs,
                'gross_profit': gross_profit,
                'gross_margin': gross_margin,
                'revenue_detail': revenue_detail,
                'cogs_detail': cogs_detail
            }
        else:
            # Get summary only
            revenue_list = "', '".join(revenue_accounts)
            cogs_list = "', '".join(cogs_accounts)
            
            query = f"""
            SELECT 
                -SUM(CASE WHEN AccountNo IN ('{revenue_list}') THEN Amount ELSE 0 END) as revenue,
                SUM(CASE WHEN AccountNo IN ('{cogs_list}') THEN Amount ELSE 0 END) as cogs
            FROM ben002.GLDetail
            WHERE EffectiveDate >= %s 
              AND EffectiveDate <= %s
              AND Posted = 1
              AND AccountNo IN ('{account_list}')
            """
            
            results = sql_service.execute_query(query, [start_date, end_date])
            
            if results and len(results) > 0:
                row = results[0]
                revenue = float(row['revenue'] or 0)
                cogs = float(row['cogs'] or 0)
                gross_profit = revenue - cogs
                gross_margin = (gross_profit / revenue * 100) if revenue > 0 else 0
                
                return {
                    'dept_code': dept_config['dept_code'],
                    'dept_name': dept_config['dept_name'],
                    'revenue': revenue,
                    'cogs': cogs,
                    'gross_profit': gross_profit,
                    'gross_margin': gross_margin
                }
        
        return {
            'dept_code': dept_config['dept_code'],
            'dept_name': dept_config['dept_name'],
            'revenue': 0,
            'cogs': 0,
            'gross_profit': 0,
            'gross_margin': 0
        }
        
    except Exception as e:
        logger.error(f"Error fetching P&L for {dept_key}: {str(e)}")
        return None


def get_other_income(start_date, end_date):
    """
    Get other income/contra-revenue from 7xxxxx accounts

    These accounts include:
    - 701000: Gain/Loss on Sale of Asset
    - 702000: Miscellaneous Income
    - 703000: A/R Discounts Allowed (contra-revenue, reduces sales)
    - 704000: A/P Discounts Taken
    - 705000: Parts Discounts (contra-revenue, reduces sales)

    In the GL, these are stored as credits (negative) for income items and
    debits (positive) for contra-revenue items. Softbase displays the net
    effect on Total Sales directly, so we use SUM(Amount) to match.

    Args:
        start_date: Start date for the report
        end_date: End date for the report

    Returns:
        Total other income (negative value reduces revenue, positive increases)
    """
    try:
        account_list = "', '".join(OTHER_INCOME_ACCOUNTS)

        # Use -SUM(Amount) to flip the sign for proper revenue calculation
        # GL stores 7xxxxx accounts as debits (positive), but they should reduce revenue
        query = f"""
        SELECT
            -SUM(Amount) as total
        FROM ben002.GLDetail
        WHERE EffectiveDate >= %s
          AND EffectiveDate <= %s
          AND Posted = 1
          AND AccountNo IN ('{account_list}')
        """
        
        results = sql_service.execute_query(query, [start_date, end_date])
        
        if results and len(results) > 0:
            return float(results[0]['total'] or 0)
        return 0
        
    except Exception as e:
        logger.error(f"Error fetching other income: {str(e)}")
        return 0


def is_full_calendar_month(start_date, end_date):
    """
    Check if the date range represents a full calendar month
    Excludes the current month to ensure GL.MTD is only used for closed/completed months
    
    Args:
        start_date: Start date string (YYYY-MM-DD)
        end_date: End date string (YYYY-MM-DD)
    
    Returns:
        Tuple of (is_full_month, year, month) or (False, None, None)
    """
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        now = datetime.now()
        
        # Check if start is first day of month
        if start.day != 1:
            return False, None, None
        
        # Check if end is last day of month
        last_day = calendar.monthrange(start.year, start.month)[1]
        if end.day != last_day:
            return False, None, None
        
        # Check if both dates are in the same month
        if start.year != end.year or start.month != end.month:
            return False, None, None
        
        # Exclude current month - always use GLDetail for in-process months
        if start.year == now.year and start.month == now.month:
            return False, None, None
        
        return True, start.year, start.month
    except:
        return False, None, None

def get_expense_data(start_date, end_date):
    """
    Get expense data organized by category
    Uses GL.MTD for full calendar months (exact Softbase match)
    Uses GLDetail for custom date ranges (flexibility)
    
    Args:
        start_date: Start date for the report
        end_date: End date for the report
    
    Returns:
        Dictionary with expense categories and totals
    """
    try:
        # Check if this is a full calendar month
        is_full_month, year, month = is_full_calendar_month(start_date, end_date)
        
        if is_full_month:
            # Use GL.MTD for exact Softbase match
            return get_expense_data_from_gl_mtd(year, month)
        else:
            # Use GLDetail for custom date ranges
            return get_expense_data_from_gldetail(start_date, end_date)
    except Exception as e:
        logger.error(f"Error in get_expense_data: {e}")
        raise

def get_expense_data_from_gl_mtd(year, month):
    """
    Get expense data from GL.MTD (monthly summary table)
    This matches Softbase P&L exactly for monthly reports
    """
    try:
        # Flatten all expense accounts
        all_expense_accounts = []  
        for category_accounts in EXPENSE_ACCOUNTS.values():
            all_expense_accounts.extend(category_accounts)
        
        expense_list = "', '".join(all_expense_accounts)
        
        query = f"""
        SELECT 
            AccountNo,
            MTD as total
        FROM ben002.GL
        WHERE Year = %s
          AND Month = %s
          AND AccountNo IN ('{expense_list}')
        """
        
        results = sql_service.execute_query(query, [year, month])
        
        # Debug logging
        print("="*80)
        print(f"EXPENSE QUERY (GL.MTD) for {year}-{month:02d}")
        print(f"Query returned {len(results if results else [])} accounts from GL.MTD")
        
        # Organize by category
        expense_data = {}
        total_expenses = 0
        
        for category, accounts in EXPENSE_ACCOUNTS.items():
            category_total = 0
            for row in results:
                account_no = str(row['AccountNo']).strip()
                if account_no in accounts:
                    amount = float(row['total'] or 0)
                    category_total += amount
            
            expense_data[category] = category_total
            total_expenses += category_total
        
        print(f"Total expenses from GL.MTD: ${total_expenses:,.2f}")
        print("="*80)
        
        expense_data['total_expenses'] = total_expenses
        return expense_data
        
    except Exception as e:
        logger.error(f"Error getting expense data from GL.MTD: {e}")
        raise

def get_expense_data_from_gldetail(start_date, end_date):
    """
    Get expense data from GLDetail (transaction-level detail)
    Used for custom date ranges that aren't full calendar months
    """
    try:
        # Flatten all expense accounts
        all_expense_accounts = []
        for category_accounts in EXPENSE_ACCOUNTS.values():
            all_expense_accounts.extend(category_accounts)
        
        expense_list = "', '".join(all_expense_accounts)
        
        # Use EffectiveDate and Posted = 1 to match revenue/COGS query logic
        query = f"""
        SELECT 
            AccountNo,
            SUM(Amount) as total
        FROM ben002.GLDetail
        WHERE EffectiveDate >= %s 
          AND EffectiveDate <= %s
          AND Posted = 1
          AND AccountNo IN ('{expense_list}')
        GROUP BY AccountNo
        """
        
        results = sql_service.execute_query(query, [start_date, end_date])
        
        # Debug: Print all accounts returned from query (using print to ensure it appears in logs)
        print("="*80)
        print(f"EXPENSE QUERY DEBUG for {start_date} to {end_date}")
        print(f"Query returned {len(results if results else [])} accounts with Posted=1 filter")
        
        # Print each account and amount
        query_total = 0
        if results:
            for row in results:
                acct = str(row['AccountNo']).strip()
                amt = float(row['total'] or 0)
                query_total += amt
                print(f"  Account {acct}: ${amt:,.2f}")
        
        print(f"Query Total (before category aggregation): ${query_total:,.2f}")
        print(f"Target from Softbase: $338,909.32")
        print(f"Difference: ${query_total - 338909.32:,.2f}")
        
        # Organize by category
        expense_data = {}
        total_expenses = 0
        matched_accounts = set()
        
        for category, accounts in EXPENSE_ACCOUNTS.items():
            category_total = 0
            for row in results:
                # Convert AccountNo to string for consistent comparison
                account_no = str(row['AccountNo']).strip()
                if account_no in accounts:
                    # Use total (posted + unposted) to match Softbase
                    amount = float(row['total'] or 0)
                    category_total += amount
                    matched_accounts.add(account_no)
            
            expense_data[category] = category_total
            total_expenses += category_total
        
        # Debug: Check for unmatched accounts
        all_result_accounts = {str(row['AccountNo']).strip() for row in results}
        unmatched = all_result_accounts - matched_accounts
        if unmatched:
            logger.warning(f"Unmatched expense accounts: {unmatched}")
            # Add unmatched amounts to total
            for row in results:
                account_no = str(row['AccountNo']).strip()
                if account_no in unmatched:
                    amount = float(row['total'] or 0)
                    total_expenses += amount
                    logger.warning(f"  Adding unmatched {account_no}: ${amount:,.2f}")
        
        expense_data['total_expenses'] = total_expenses
        
        print(f"After category aggregation: ${total_expenses:,.2f}")
        print("Note: Using GLDetail for custom date range (not full month)")
        print("="*80)
        
        return expense_data
        
    except Exception as e:
        logger.error(f"Error fetching expense data: {str(e)}")
        return {'total_expenses': 0}


@pl_report_bp.route('/api/reports/pl', methods=['GET'])
def get_pl_report():
    """
    Get P&L report with departmental and consolidated views
    
    Query Parameters:
        start_date: Start date (YYYY-MM-DD)
        end_date: End date (YYYY-MM-DD)
        view: 'consolidated' or 'departmental' (default: 'consolidated')
    
    Returns:
        JSON with P&L data
    """
    try:
        # Get query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        view = request.args.get('view', 'consolidated')
        
        if not start_date or not end_date:
            return jsonify({
                'success': False,
                'error': 'start_date and end_date are required'
            }), 400
        
        # Validate dates
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }), 400
        
        # Check if detail is requested
        include_detail = request.args.get('detail', 'false').lower() == 'true'
        
        # Get data for all departments
        departments = {}
        total_revenue = 0
        total_cogs = 0
        
        for dept_key in GL_ACCOUNTS.keys():
            dept_data = get_department_data(start_date, end_date, dept_key, include_detail)
            if dept_data:
                # Include all departments in totals
                total_revenue += dept_data['revenue']
                total_cogs += dept_data['cogs']
                
                # But exclude administrative from the department breakdown display
                if dept_key != 'administrative':
                    departments[dept_key] = dept_data
        
        # Add other income (7xxxxx accounts - contra-revenue/other income)
        other_income = get_other_income(start_date, end_date)
        total_revenue += other_income
        
        # Get expense data
        expenses = get_expense_data(start_date, end_date)
        
        # Calculate consolidated metrics
        total_gross_profit = total_revenue - total_cogs
        gross_margin = (total_gross_profit / total_revenue * 100) if total_revenue > 0 else 0
        operating_profit = total_gross_profit - expenses['total_expenses']
        operating_margin = (operating_profit / total_revenue * 100) if total_revenue > 0 else 0
        
        # Build response
        response = {
            'success': True,
            'start_date': start_date,
            'end_date': end_date,
            'consolidated': {
                'revenue': total_revenue,
                'cogs': total_cogs,
                'gross_profit': total_gross_profit,
                'gross_margin': gross_margin,
                'operating_expenses': expenses['total_expenses'],
                'operating_profit': operating_profit,
                'operating_margin': operating_margin
            },
            'departments': departments,
            'expenses': expenses
        }
        
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"Error generating P&L report: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@pl_report_bp.route('/api/reports/pl/mtd', methods=['GET'])
def get_pl_mtd():
    """
    Get Month-to-Date P&L report
    
    Query Parameters:
        year: Year (default: current year)
        month: Month (default: current month)
    
    Returns:
        JSON with MTD P&L data
    """
    try:
        # Get current date
        now = datetime.now()
        year = request.args.get('year', now.year, type=int)
        month = request.args.get('month', now.month, type=int)
        
        # Calculate MTD date range
        start_date = f"{year}-{month:02d}-01"
        
        # Get last day of month
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        
        from datetime import timedelta
        last_day = next_month - timedelta(days=1)
        end_date = last_day.strftime('%Y-%m-%d')
        
        # Forward to main P&L endpoint
        request.args = {'start_date': start_date, 'end_date': end_date}
        return get_pl_report()
        
    except Exception as e:
        logger.error(f"Error generating MTD P&L: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@pl_report_bp.route('/api/reports/pl/ytd', methods=['GET'])
def get_pl_ytd():
    """
    Get Fiscal Year-to-Date P&L report
    
    Query Parameters:
        year: Year (optional, defaults to current fiscal year)
    
    Returns:
        JSON with fiscal YTD P&L data
    """
    try:
        # Get current date
        now = datetime.now()
        
        # Calculate fiscal YTD date range
        fiscal_ytd_start = get_fiscal_ytd_start()
        start_date = fiscal_ytd_start.strftime('%Y-%m-%d')
        end_date = now.strftime('%Y-%m-%d')
        
        # Forward to main P&L endpoint
        request.args = {'start_date': start_date, 'end_date': end_date}
        return get_pl_report()
        
    except Exception as e:
        logger.error(f"Error generating YTD P&L: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@pl_report_bp.route('/api/reports/pl/export-excel', methods=['GET'])
def export_pl_excel():
    """
    Export P&L report to Excel with template format
    Replicates the structure from October 2025 Prelim Financials.xlsx
    
    Query Parameters:
        month: Month number (1-12), defaults to current month
        year: Year (YYYY), defaults to current year
    
    Returns:
        Excel file with Profit & Loss Consolidated sheet (and more sheets in future phases)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from flask import send_file
        import io
        from datetime import datetime
        import calendar
        
        # Get month and year from query parameters
        now = datetime.now()
        month = request.args.get('month', now.month, type=int)
        year = request.args.get('year', now.year, type=int)
        
        if not (1 <= month <= 12):
            return jsonify({'error': 'month must be between 1 and 12'}), 400
        
        # Calculate MTD date range
        first_day = f"{year}-{month:02d}-01"
        last_day_num = calendar.monthrange(year, month)[1]
        last_day = f"{year}-{month:02d}-{last_day_num:02d}"
        
        # Calculate fiscal YTD date range (fiscal year starts in November)
        if month >= 11:
            ytd_start = f"{year}-11-01"
        else:
            ytd_start = f"{year-1}-11-01"
        ytd_end = last_day
        
        logger.info(f"Generating P&L Excel for {calendar.month_name[month]} {year}")
        logger.info(f"MTD: {first_day} to {last_day}")
        logger.info(f"YTD: {ytd_start} to {ytd_end}")
        
        # Get MTD and YTD data for all departments
        mtd_data = get_all_departments_data(first_day, last_day)
        ytd_data = get_all_departments_data(ytd_start, ytd_end)
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create Profit & Loss Consolidated sheet
        ws = wb.create_sheet("Profit & Loss Consolidated")
        
        # Define styles
        header_font = Font(bold=True, size=11)
        bold_font = Font(bold=True)
        currency_format = '$#,##0.00'
        percent_format = '0.0%'
        
        # Row 1: Title
        ws['B1'] = f"Profit & Loss Statement - {calendar.month_name[month]} {year}"
        ws['B1'].font = Font(bold=True, size=14)
        
        # Row 2: Company Name
        ws['B2'] = "Bennett Material Handling"
        ws['B2'].font = Font(bold=True, size=12)
        
        # Row 3: Department Headers
        ws['B3'] = "Bennett Material Handling"
        ws['B3'].font = header_font
        
        dept_headers = [
            ("C3", "New Equipment", 10),
            ("D3", "Used Equipment", 20),
            ("E3", "Parts", 30),
            ("F3", "Field Service", 40),
            ("G3", "Rental", 60),
            ("H3", "Transportation / Trucking", 80),
            ("I3", "In House / Administrative", 90),
            ("J3", "Total", None)
        ]
        
        for cell, header, dept_code in dept_headers:
            ws[cell] = header
            ws[cell].font = header_font
        
        # Row 4: Department codes
        for cell, _, dept_code in dept_headers:
            if dept_code:
                row = int(cell[1:])
                ws[cell.replace('3', '4')] = dept_code
                ws[cell.replace('3', '4')].font = header_font
        
        # MTD Section (Rows 6-15)
        ws['B5'] = "MTD (Month-to-Date)"
        ws['B5'].font = Font(bold=True, size=11)
        
        # Row 6: Income
        ws['B6'] = "Income"
        row_num = 6
        for idx, (dept_key, dept_config) in enumerate(GL_ACCOUNTS.items(), start=3):
            col = get_column_letter(idx)
            revenue = mtd_data[dept_key]['revenue']
            ws[f'{col}{row_num}'] = revenue
            ws[f'{col}{row_num}'].number_format = currency_format
        
        # Total column
        ws[f'J{row_num}'] = f'=SUM(C{row_num}:I{row_num})'
        ws[f'J{row_num}'].number_format = currency_format
        
        # Row 7: Cost of Goods Sold
        ws['B7'] = "Cost of Goods Sold"
        row_num = 7
        for idx, (dept_key, dept_config) in enumerate(GL_ACCOUNTS.items(), start=3):
            col = get_column_letter(idx)
            cogs = mtd_data[dept_key]['cogs']
            ws[f'{col}{row_num}'] = cogs
            ws[f'{col}{row_num}'].number_format = currency_format
        
        ws[f'J{row_num}'] = f'=SUM(C{row_num}:I{row_num})'
        ws[f'J{row_num}'].number_format = currency_format
        
        # Row 8: Gross Profit
        ws['B8'] = "Gross Profit"
        row_num = 8
        for col_idx in range(3, 11):  # C to J
            col = get_column_letter(col_idx)
            ws[f'{col}{row_num}'] = f'={col}6-{col}7'
            ws[f'{col}{row_num}'].number_format = currency_format
        
        # Row 9: Gross Margin (BOLD)
        ws['B9'] = "Gross Margin"
        ws['B9'].font = bold_font
        row_num = 9
        for col_idx in range(3, 11):
            col = get_column_letter(col_idx)
            ws[f'{col}{row_num}'] = f'=IF({col}6<>0,{col}8/{col}6,0)'
            ws[f'{col}{row_num}'].number_format = percent_format
            ws[f'{col}{row_num}'].font = bold_font
        
        # Row 10: Overhead Expenses
        ws['B10'] = "Overhead Expenses"
        row_num = 10
        overhead = get_overhead_expenses(first_day, last_day)
        # Distribute overhead across departments (simplified - all in Admin column I)
        ws['I10'] = overhead
        ws['I10'].number_format = currency_format
        ws['J10'] = f'=SUM(C10:I10)'
        ws['J10'].number_format = currency_format
        
        # Row 11: Operating Profit
        ws['B11'] = "Operating Profit"
        row_num = 11
        for col_idx in range(3, 11):
            col = get_column_letter(col_idx)
            ws[f'{col}{row_num}'] = f'={col}8-{col}10'
            ws[f'{col}{row_num}'].number_format = currency_format
        
        # Row 12: Operating Margin (BOLD)
        ws['B12'] = "Operating  Margin"
        ws['B12'].font = bold_font
        ws['J12'] = '=IF(J6<>0,J11/J6,0)'
        ws['J12'].number_format = percent_format
        ws['J12'].font = bold_font
        
        # Row 13: Other Income & Expense
        ws['B13'] = "Other Income & Expense"
        row_num = 13
        other_income = get_other_income(first_day, last_day)
        ws['I13'] = other_income
        ws['I13'].number_format = currency_format
        ws['J13'] = f'=SUM(C13:I13)'
        ws['J13'].number_format = currency_format
        
        # Row 14: Net Profit (BOLD)
        ws['B14'] = "Net Profit"
        ws['B14'].font = bold_font
        row_num = 14
        for col_idx in range(3, 11):
            col = get_column_letter(col_idx)
            ws[f'{col}{row_num}'] = f'={col}11-{col}13'
            ws[f'{col}{row_num}'].number_format = currency_format
            ws[f'{col}{row_num}'].font = bold_font
        
        # Row 15: Net Margin (BOLD)
        ws['B15'] = "Net  Margin"
        ws['B15'].font = bold_font
        ws['J15'] = '=IF(J6<>0,J14/J6,0)'
        ws['J15'].number_format = percent_format
        ws['J15'].font = bold_font
        
        # YTD Section (Rows 18-29) - Same structure as MTD
        ws['B18'] = "YTD Summary"
        ws['B18'].font = Font(bold=True, size=11)
        
        # Row 19: Department Headers (repeat)
        ws['B19'] = "Bennett Material Handling"
        ws['B19'].font = header_font
        for cell, header, dept_code in dept_headers:
            row = int(cell[1:])
            new_cell = cell.replace('3', '19')
            ws[new_cell] = header
            ws[new_cell].font = header_font
        
        # Row 20: Income (YTD)
        ws['B20'] = "Income"
        row_num = 20
        for idx, (dept_key, dept_config) in enumerate(GL_ACCOUNTS.items(), start=3):
            col = get_column_letter(idx)
            revenue = ytd_data[dept_key]['revenue']
            ws[f'{col}{row_num}'] = revenue
            ws[f'{col}{row_num}'].number_format = currency_format
        
        ws[f'J{row_num}'] = f'=SUM(C{row_num}:I{row_num})'
        ws[f'J{row_num}'].number_format = currency_format
        
        # Row 21: COGS (YTD)
        ws['B21'] = "Cost of Goods Sold"
        row_num = 21
        for idx, (dept_key, dept_config) in enumerate(GL_ACCOUNTS.items(), start=3):
            col = get_column_letter(idx)
            cogs = ytd_data[dept_key]['cogs']
            ws[f'{col}{row_num}'] = cogs
            ws[f'{col}{row_num}'].number_format = currency_format
        
        ws[f'J{row_num}'] = f'=SUM(C{row_num}:I{row_num})'
        ws[f'J{row_num}'].number_format = currency_format
        
        # Row 22: Gross Profit (YTD)
        ws['B22'] = "Gross Profit"
        row_num = 22
        for col_idx in range(3, 11):
            col = get_column_letter(col_idx)
            ws[f'{col}{row_num}'] = f'={col}20-{col}21'
            ws[f'{col}{row_num}'].number_format = currency_format
        
        # Row 23: Gross Margin (YTD, BOLD)
        ws['B23'] = "Gross Margin"
        ws['B23'].font = bold_font
        row_num = 23
        for col_idx in range(3, 11):
            col = get_column_letter(col_idx)
            ws[f'{col}{row_num}'] = f'=IF({col}20<>0,{col}22/{col}20,0)'
            ws[f'{col}{row_num}'].number_format = percent_format
            ws[f'{col}{row_num}'].font = bold_font
        
        # Row 24: Overhead (YTD)
        ws['B24'] = "Overhead Expenses"
        row_num = 24
        overhead_ytd = get_overhead_expenses(ytd_start, ytd_end)
        ws['I24'] = overhead_ytd
        ws['I24'].number_format = currency_format
        ws['J24'] = f'=SUM(C24:I24)'
        ws['J24'].number_format = currency_format
        
        # Row 25: Operating Profit (YTD)
        ws['B25'] = "Operating Profit"
        row_num = 25
        for col_idx in range(3, 11):
            col = get_column_letter(col_idx)
            ws[f'{col}{row_num}'] = f'={col}22-{col}24'
            ws[f'{col}{row_num}'].number_format = currency_format
        
        # Row 26: Operating Margin (YTD, BOLD)
        ws['B26'] = "Operating  Margin"
        ws['B26'].font = bold_font
        ws['J26'] = '=IF(J20<>0,J25/J20,0)'
        ws['J26'].number_format = percent_format
        ws['J26'].font = bold_font
        
        # Row 27: Other Income (YTD)
        ws['B27'] = "Other Income & Expense"
        row_num = 27
        other_income_ytd = get_other_income(ytd_start, ytd_end)
        ws['I27'] = other_income_ytd
        ws['I27'].number_format = currency_format
        ws['J27'] = f'=SUM(C27:I27)'
        ws['J27'].number_format = currency_format
        
        # Row 28: Net Profit (YTD, BOLD)
        ws['B28'] = "Net Profit"
        ws['B28'].font = bold_font
        row_num = 28
        for col_idx in range(3, 11):
            col = get_column_letter(col_idx)
            ws[f'{col}{row_num}'] = f'={col}25-{col}27'
            ws[f'{col}{row_num}'].number_format = currency_format
            ws[f'{col}{row_num}'].font = bold_font
        
        # Row 29: Net Margin (YTD, BOLD)
        ws['B29'] = "Net  Margin"
        ws['B29'].font = bold_font
        ws['J29'] = '=IF(J20<>0,J28/J20,0)'
        ws['J29'].number_format = percent_format
        ws['J29'].font = bold_font
        
        # Set column widths
        ws.column_dimensions['B'].width = 30
        for col_idx in range(3, 11):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"ProfitLossReport_{year}_{month:02d}.xlsx"
        
        logger.info(f"P&L Excel export generated: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting P&L Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to export P&L to Excel', 'message': str(e)}), 500


def get_all_departments_data(start_date, end_date):
    """Get revenue and COGS for all departments"""
    dept_data = {}
    
    for dept_key, dept_config in GL_ACCOUNTS.items():
        revenue_accounts = dept_config['revenue']
        cogs_accounts = dept_config['cogs']
        
        # Get revenue (negative because credits)
        revenue_query = f"""
        SELECT -SUM(Amount) as total
        FROM ben002.GLDetail
        WHERE AccountNo IN ('{"', '".join(revenue_accounts)}')
          AND EffectiveDate >= %s
          AND EffectiveDate <= %s
          AND Posted = 1
        """
        
        revenue_result = sql_service.execute_query(revenue_query, [start_date, end_date])
        revenue = float(revenue_result[0]['total'] or 0) if revenue_result else 0
        
        # Get COGS (positive because debits)
        cogs_query = f"""
        SELECT SUM(Amount) as total
        FROM ben002.GLDetail
        WHERE AccountNo IN ('{"', '".join(cogs_accounts)}')
          AND EffectiveDate >= %s
          AND EffectiveDate <= %s
          AND Posted = 1
        """
        
        cogs_result = sql_service.execute_query(cogs_query, [start_date, end_date])
        cogs = float(cogs_result[0]['total'] or 0) if cogs_result else 0
        
        dept_data[dept_key] = {
            'revenue': revenue,
            'cogs': cogs,
            'gross_profit': revenue - cogs
        }
    
    return dept_data


def get_overhead_expenses(start_date, end_date):
    """Get total overhead expenses (6xxxx accounts)"""
    all_expense_accounts = []
    for category, accounts in EXPENSE_ACCOUNTS.items():
        all_expense_accounts.extend(accounts)
    
    query = f"""
    SELECT SUM(Amount) as total
    FROM ben002.GLDetail
    WHERE AccountNo IN ('{"', '".join(all_expense_accounts)}')
      AND EffectiveDate >= %s
      AND EffectiveDate <= %s
      AND Posted = 1
    """
    
    result = sql_service.execute_query(query, [start_date, end_date])
    return float(result[0]['total'] or 0) if result else 0


# Note: get_other_income is defined earlier in this file (line ~327)
# The function was previously duplicated here - removed to avoid confusion
