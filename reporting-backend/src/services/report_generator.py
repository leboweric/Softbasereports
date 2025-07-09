import pandas as pd
import io
import json
from datetime import datetime, timedelta
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
from io import BytesIO
import base64

class ReportGenerator:
    def __init__(self):
        self.setup_matplotlib()
    
    def setup_matplotlib(self):
        """Configure matplotlib for better chart generation"""
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")
    
    def generate_csv_report(self, data, filename="report.csv"):
        """Generate CSV report from data"""
        try:
            df = pd.DataFrame(data)
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False)
            return csv_buffer.getvalue()
        except Exception as e:
            raise Exception(f"Error generating CSV: {str(e)}")
    
    def generate_excel_report(self, data, filename="report.xlsx", chart_type=None):
        """Generate Excel report with formatting and optional charts"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Report Data"
            
            if not data:
                return None
            
            df = pd.DataFrame(data)
            
            # Add headers with formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add chart if requested
            if chart_type and len(df) > 1:
                self._add_excel_chart(ws, df, chart_type)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer.getvalue()
            
        except Exception as e:
            raise Exception(f"Error generating Excel: {str(e)}")
    
    def _add_excel_chart(self, worksheet, df, chart_type):
        """Add chart to Excel worksheet"""
        try:
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            else:
                return
            
            # Assume first column is categories, second is values
            if len(df.columns) >= 2:
                data = Reference(worksheet, min_col=2, min_row=1, max_row=len(df)+1, max_col=2)
                cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(df)+1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.title = "Report Chart"
                chart.x_axis.title = df.columns[0]
                chart.y_axis.title = df.columns[1]
                
                worksheet.add_chart(chart, "E2")
        except Exception as e:
            print(f"Warning: Could not add chart: {str(e)}")
    
    def generate_pdf_report(self, data, title="Report", filename="report.pdf"):
        """Generate PDF report with formatting"""
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            
            # Title
            pdf.cell(0, 10, title, ln=True, align="C")
            pdf.ln(10)
            
            # Date
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 10, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
            pdf.ln(5)
            
            if not data:
                pdf.set_font("Arial", "", 12)
                pdf.cell(0, 10, "No data available", ln=True)
                return pdf.output(dest='S').encode('latin-1')
            
            df = pd.DataFrame(data)
            
            # Table headers
            pdf.set_font("Arial", "B", 10)
            col_width = 190 / len(df.columns)
            
            for column in df.columns:
                pdf.cell(col_width, 10, str(column), border=1, align="C")
            pdf.ln()
            
            # Table data
            pdf.set_font("Arial", "", 9)
            for _, row in df.iterrows():
                for value in row:
                    # Truncate long values
                    cell_value = str(value)[:20] + "..." if len(str(value)) > 20 else str(value)
                    pdf.cell(col_width, 8, cell_value, border=1, align="C")
                pdf.ln()
            
            return pdf.output(dest='S').encode('latin-1')
            
        except Exception as e:
            raise Exception(f"Error generating PDF: {str(e)}")
    
    def generate_chart_image(self, data, chart_type="bar", title="Chart", x_label="X", y_label="Y"):
        """Generate chart image as base64 string"""
        try:
            if not data:
                return None
            
            df = pd.DataFrame(data)
            
            plt.figure(figsize=(10, 6))
            
            if chart_type == "bar" and len(df.columns) >= 2:
                plt.bar(df.iloc[:, 0], df.iloc[:, 1])
            elif chart_type == "line" and len(df.columns) >= 2:
                plt.plot(df.iloc[:, 0], df.iloc[:, 1], marker='o')
            elif chart_type == "pie" and len(df.columns) >= 2:
                plt.pie(df.iloc[:, 1], labels=df.iloc[:, 0], autopct='%1.1f%%')
            else:
                # Default to simple bar chart
                if len(df.columns) >= 2:
                    plt.bar(range(len(df)), df.iloc[:, 1])
                    plt.xticks(range(len(df)), df.iloc[:, 0], rotation=45)
            
            plt.title(title)
            plt.xlabel(x_label)
            plt.ylabel(y_label)
            plt.tight_layout()
            
            # Convert to base64
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
            img_buffer.seek(0)
            img_base64 = base64.b64encode(img_buffer.getvalue()).decode()
            plt.close()
            
            return img_base64
            
        except Exception as e:
            raise Exception(f"Error generating chart: {str(e)}")
    
    def generate_dashboard_metrics(self, data):
        """Generate key metrics for dashboard"""
        try:
            if not data:
                return {}
            
            df = pd.DataFrame(data)
            
            metrics = {
                "total_records": len(df),
                "date_range": {
                    "start": None,
                    "end": None
                }
            }
            
            # Try to find date columns
            date_columns = []
            for col in df.columns:
                if 'date' in col.lower() or 'time' in col.lower():
                    try:
                        df[col] = pd.to_datetime(df[col])
                        date_columns.append(col)
                    except:
                        pass
            
            if date_columns:
                date_col = date_columns[0]
                metrics["date_range"]["start"] = df[date_col].min().isoformat() if pd.notna(df[date_col].min()) else None
                metrics["date_range"]["end"] = df[date_col].max().isoformat() if pd.notna(df[date_col].max()) else None
            
            # Try to find numeric columns for aggregation
            numeric_columns = df.select_dtypes(include=['number']).columns
            if len(numeric_columns) > 0:
                metrics["numeric_summary"] = {}
                for col in numeric_columns:
                    metrics["numeric_summary"][col] = {
                        "sum": float(df[col].sum()),
                        "avg": float(df[col].mean()),
                        "min": float(df[col].min()),
                        "max": float(df[col].max())
                    }
            
            return metrics
            
        except Exception as e:
            return {"error": f"Error generating metrics: {str(e)}"}
    
    def generate_trend_analysis(self, data, date_column=None, value_column=None):
        """Generate trend analysis for time series data"""
        try:
            if not data:
                return {}
            
            df = pd.DataFrame(data)
            
            # Auto-detect date and value columns if not specified
            if not date_column:
                for col in df.columns:
                    if 'date' in col.lower() or 'time' in col.lower():
                        date_column = col
                        break
            
            if not value_column:
                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    value_column = numeric_cols[0]
            
            if not date_column or not value_column:
                return {"error": "Could not identify date and value columns"}
            
            # Convert date column
            df[date_column] = pd.to_datetime(df[date_column])
            df = df.sort_values(date_column)
            
            # Calculate trends
            trend_data = {
                "period_over_period": {},
                "moving_averages": {},
                "growth_rate": 0
            }
            
            # Calculate period-over-period changes
            if len(df) > 1:
                latest_value = df[value_column].iloc[-1]
                previous_value = df[value_column].iloc[-2]
                if previous_value != 0:
                    trend_data["period_over_period"]["change_percent"] = ((latest_value - previous_value) / previous_value) * 100
                else:
                    trend_data["period_over_period"]["change_percent"] = 0
                
                trend_data["period_over_period"]["change_absolute"] = latest_value - previous_value
            
            # Calculate moving averages
            if len(df) >= 7:
                df['ma_7'] = df[value_column].rolling(window=7).mean()
                trend_data["moving_averages"]["7_day"] = df['ma_7'].iloc[-1] if pd.notna(df['ma_7'].iloc[-1]) else None
            
            if len(df) >= 30:
                df['ma_30'] = df[value_column].rolling(window=30).mean()
                trend_data["moving_averages"]["30_day"] = df['ma_30'].iloc[-1] if pd.notna(df['ma_30'].iloc[-1]) else None
            
            # Calculate overall growth rate
            if len(df) > 1:
                first_value = df[value_column].iloc[0]
                last_value = df[value_column].iloc[-1]
                if first_value != 0:
                    days_diff = (df[date_column].iloc[-1] - df[date_column].iloc[0]).days
                    if days_diff > 0:
                        trend_data["growth_rate"] = ((last_value / first_value) ** (365 / days_diff) - 1) * 100
            
            return trend_data
            
        except Exception as e:
            return {"error": f"Error generating trend analysis: {str(e)}"}

